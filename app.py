#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import io
import base64
import sqlite3
import hashlib
import secrets
import smtplib
import time
from werkzeug.middleware.proxy_fix import ProxyFix
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from datetime import datetime, date, timedelta
from flask import Flask, request, redirect, url_for, session, abort, send_file, Response, flash
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "uploads", "it")
ANN_UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "uploads", "announcements")
os.makedirs(ANN_UPLOAD_DIR, exist_ok=True)
ALLOWED_IMAGE_EXT = {"png", "jpg", "jpeg", "webp"}

def allowed_image(filename):
    if not filename:
        return False
    if "." not in filename:
        return False
    return filename.rsplit(".", 1)[1].lower() in ALLOWED_IMAGE_EXT


def public_base_url():
    host = request.host
    return f"https://{host}"


import qrcode
from openpyxl import Workbook


# -----------------------------
# CONFIG
# -----------------------------
APP_TITLE = "İK Portal"
DB_PATH = os.environ.get("IK_DB_PATH", os.path.join(os.path.dirname(__file__), "ik.db"))
SECRET_KEY = os.environ.get("IK_SECRET_KEY", "saDfasdwqefgdsvcxzfasdRqweasDASFGA")

# SMTP (systemd env ile ver; yoksa mail sessizce pas geçer)
SMTP_HOST = os.environ.get("IK_SMTP_HOST", "")
SMTP_PORT = int(os.environ.get("IK_SMTP_PORT", "587"))
SMTP_USER = os.environ.get("IK_SMTP_USER", "")
SMTP_PASS = os.environ.get("IK_SMTP_PASS", "")
MAIL_FROM  = os.environ.get("IK_MAIL_FROM", SMTP_USER)

QR_COOLDOWN_MINUTES = int(os.environ.get("IK_QR_COOLDOWN_MINUTES", "10"))

# Roles
ROLE_OWNER = "owner"
ROLE_ACCOUNTING = "accounting"
ROLE_MANAGER = "manager"
ROLE_PERSONNEL = "personnel"
ROLE_RESPONSIBLE = "responsible"   # ✅ NEW
ROLE_IT = "it"
ALLOWED_ROLES = {ROLE_OWNER, ROLE_ACCOUNTING, ROLE_MANAGER, ROLE_PERSONNEL, ROLE_RESPONSIBLE, ROLE_IT}

# Attendance types
TUR_GIRIS = "GIRIS"
TUR_OGLE_CIKIS = "OGLE_CIKIS"
TUR_OGLE_GIRIS = "OGLE_GIRIS"
TUR_CIKIS = "CIKIS"
TUR_SEQUENCE = [TUR_GIRIS, TUR_OGLE_CIKIS, TUR_OGLE_GIRIS, TUR_CIKIS]
TUR_LABEL = {
    TUR_GIRIS: "Giriş",
    TUR_OGLE_CIKIS: "Öğle Çıkış",
    TUR_OGLE_GIRIS: "Öğle Giriş",
    TUR_CIKIS: "Çıkış",
}

# Leave statuses
STATUS_PENDING = "Beklemede"
STATUS_APPROVED = "Onaylandi"
STATUS_REJECTED = "Reddedildi"

# Advance statuses (✅ NEW)
ADV_STATUS_PENDING = "Beklemede"
ADV_STATUS_SENT_TO_OWNER = "PatronOnayinda"
ADV_STATUS_APPROVED = "Onaylandi"
ADV_STATUS_REJECTED = "Reddedildi"

# Leave stages (✅ NEW for multi-step approval)
STAGE_RESPONSIBLE = "RESPONSIBLE"
STAGE_MANAGER = "MANAGER"
STAGE_OWNER = "OWNER"
STAGE_DONE = "DONE"


app = Flask(__name__)
app.secret_key = SECRET_KEY

# Nginx reverse proxy arkasında https/http doğru algılansın
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)

# -----------------------------
# DB helpers
# -----------------------------
def db_connect():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def query_one(sql, params=()):
    conn = db_connect()
    cur = conn.execute(sql, params)
    row = cur.fetchone()
    conn.close()
    return row

def query_all(sql, params=()):
    conn = db_connect()
    cur = conn.execute(sql, params)
    rows = cur.fetchall()
    conn.close()
    return rows

def exec_sql(sql, params=()):
    conn = db_connect()
    cur = conn.execute(sql, params)
    conn.commit()
    last_id = cur.lastrowid
    conn.close()
    return last_id

def is_demo_mode():
    return os.environ.get("IK_DEMO_MODE", "0") == "1"


def is_protected_user(row_or_uid):
    """
    Demo modda düzenlenmesi/silinmesi yasak kullanıcıyı tespit eder.
    """

    if row_or_uid is None:
        return False

    uid = None
    role = ""

    # sqlite3.Row gelirse
    if isinstance(row_or_uid, sqlite3.Row):
        uid = row_or_uid["id"]
        role = (row_or_uid["role"] or "").strip()

    # dict gelirse
    elif isinstance(row_or_uid, dict):
        uid = row_or_uid.get("id")
        role = (row_or_uid.get("role") or "").strip()

    # direkt id gelirse
    else:
        try:
            uid = int(row_or_uid)
        except Exception:
            uid = None

    # uid güvenli int'e çevir
    try:
        uid = int(uid)
    except Exception:
        uid = None

    # role boşsa DB'den çek
    if (not role) and uid is not None:
        try:
            r = query_one("SELECT role FROM users WHERE id=?", (uid,))
            if r:
                role = (r["role"] or "").strip()
        except Exception:
            role = ""

    # ✅ ANA ADMIN (id=1) HER ZAMAN KORUNUR
    if uid == 1:
        return True

    # ✅ Owner rolü olan da korunur
    if role.lower() == "owner":
        return True

    return False

def table_has_column(conn, table, col):
    cur = conn.execute(f"PRAGMA table_info({table})")
    cols = [r[1] for r in cur.fetchall()]
    return col in cols

def init_db():
    conn = db_connect()

    # Users:
    # - email: mail atmak için
    # - manager_id: personelin müdürü (users.id)
    # - responsible_id: personelin sorumlusu (users.id) ✅ NEW
    # - annual_leave_days: yıllık izin hakkı
    # - hire_date: işe giriş tarihi (opsiyon)
    conn.execute("""
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        full_name TEXT NOT NULL,
        username TEXT NOT NULL UNIQUE,
        password_hash TEXT NOT NULL,
        role TEXT NOT NULL DEFAULT 'personnel',
        is_active INTEGER NOT NULL DEFAULT 1,

        email TEXT,
        manager_id INTEGER,
        responsible_id INTEGER,

        hire_date TEXT,
        annual_leave_days INTEGER NOT NULL DEFAULT 15,
        annual_leave_override INTEGER,

        can_qr INTEGER NOT NULL DEFAULT 0,
        qr_secret TEXT,

        FOREIGN KEY(manager_id) REFERENCES users(id),
        FOREIGN KEY(responsible_id) REFERENCES users(id)
    )
    """)

    # Leave requests
    conn.execute("""
    CREATE TABLE IF NOT EXISTS leave_requests (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        start_date TEXT NOT NULL,
        end_date TEXT NOT NULL,
        leave_type TEXT NOT NULL DEFAULT 'Yillik',
        reason TEXT,
        is_half_day INTEGER NOT NULL DEFAULT 0,
        half_day_part TEXT,
        status TEXT NOT NULL DEFAULT 'Beklemede',
        created_at TEXT NOT NULL,
        decided_at TEXT,
        decided_by INTEGER,

        pending_with INTEGER,
        stage TEXT,

        FOREIGN KEY(user_id) REFERENCES users(id),
        FOREIGN KEY(decided_by) REFERENCES users(id),
        FOREIGN KEY(pending_with) REFERENCES users(id)
    )
    """)

    # Attendance logs
    conn.execute("""
    CREATE TABLE IF NOT EXISTS attendance (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        tarih TEXT NOT NULL,
        saat  TEXT NOT NULL,
        tur   TEXT NOT NULL,
        ip    TEXT,
        cihaz TEXT,
        created_at TEXT NOT NULL,
        FOREIGN KEY(user_id) REFERENCES users(id)
    )
    """)

    # ✅ ZIMMET (assets) NEW
    conn.execute("""
    CREATE TABLE IF NOT EXISTS assets (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        item_name TEXT NOT NULL,
        item_props TEXT,
        serial_no TEXT,
        assigned_at TEXT NOT NULL,
        assigned_by INTEGER,
        returned_at TEXT,
        returned_by INTEGER,
        note TEXT,
        FOREIGN KEY(user_id) REFERENCES users(id),
        FOREIGN KEY(assigned_by) REFERENCES users(id),
        FOREIGN KEY(returned_by) REFERENCES users(id)
    )
    """)

    # ✅ AVANS (advances) NEW
    conn.execute("""
    CREATE TABLE IF NOT EXISTS advances (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        amount REAL NOT NULL,
        reason TEXT,
        status TEXT NOT NULL DEFAULT 'Beklemede',
        created_at TEXT NOT NULL,
        decided_at TEXT,
        decided_by INTEGER,
        forwarded_at TEXT,
        forwarded_by INTEGER,
        FOREIGN KEY(user_id) REFERENCES users(id),
        FOREIGN KEY(decided_by) REFERENCES users(id),
        FOREIGN KEY(forwarded_by) REFERENCES users(id)
    )
    """)

    # ✅ IT TALEP (tickets) NEW
    conn.execute("""
    CREATE TABLE IF NOT EXISTS it_tickets (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,

        subject TEXT NOT NULL,
        description TEXT,
        priority TEXT NOT NULL DEFAULT 'Orta',          -- Düşük/Orta/Yüksek/Acil

        status TEXT NOT NULL DEFAULT 'Açık',            -- Açık/Beklemede/Yönlendirildi/Kapalı
        status_note TEXT,                               -- IT açıklaması / durum notu

        image_path TEXT,                                -- yüklenen resim dosya yolu

        created_at TEXT NOT NULL,
        updated_at TEXT,
        closed_at TEXT,

        FOREIGN KEY(user_id) REFERENCES users(id)
    )
    """)

    # ✅ IT TALEP MESAJLAR (ticket_messages) NEW
    conn.execute("""
    CREATE TABLE IF NOT EXISTS it_ticket_messages (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        ticket_id INTEGER NOT NULL,
        author_id INTEGER NOT NULL,                     -- mesajı yazan (IT veya talep sahibi)
        message TEXT NOT NULL,
        created_at TEXT NOT NULL,

        FOREIGN KEY(ticket_id) REFERENCES it_tickets(id),
        FOREIGN KEY(author_id) REFERENCES users(id)
    )
    """)


    # Migrations (older db -> add missing columns safely)
    for col, ddl in [
        ("email", "ALTER TABLE users ADD COLUMN email TEXT"),
        ("manager_id", "ALTER TABLE users ADD COLUMN manager_id INTEGER"),
        ("responsible_id", "ALTER TABLE users ADD COLUMN responsible_id INTEGER"),  # ✅ NEW
        ("hire_date", "ALTER TABLE users ADD COLUMN hire_date TEXT"),
        ("annual_leave_days", "ALTER TABLE users ADD COLUMN annual_leave_days INTEGER NOT NULL DEFAULT 15"),
        ("can_qr", "ALTER TABLE users ADD COLUMN can_qr INTEGER NOT NULL DEFAULT 0"),
        ("qr_secret", "ALTER TABLE users ADD COLUMN qr_secret TEXT"),
        ("annual_leave_override", "ALTER TABLE users ADD COLUMN annual_leave_override INTEGER"),
        ("ann_dismissed_at", "ALTER TABLE users ADD COLUMN ann_dismissed_at TEXT"),
    ]:
        if not table_has_column(conn, "users", col):
            conn.execute(ddl)

    for col, ddl in [
        ("decided_at", "ALTER TABLE leave_requests ADD COLUMN decided_at TEXT"),
        ("decided_by", "ALTER TABLE leave_requests ADD COLUMN decided_by INTEGER"),
        ("pending_with", "ALTER TABLE leave_requests ADD COLUMN pending_with INTEGER"),  # ✅ NEW
        ("stage", "ALTER TABLE leave_requests ADD COLUMN stage TEXT"),                  # ✅ NEW
        ("is_half_day", "ALTER TABLE leave_requests ADD COLUMN is_half_day INTEGER NOT NULL DEFAULT 0"),
        ("half_day_part", "ALTER TABLE leave_requests ADD COLUMN half_day_part TEXT"),
    ]:
        if not table_has_column(conn, "leave_requests", col):
            conn.execute(ddl)

    # -----------------------------
    # ✅ DUYURULAR (announcements) NEW
    # -----------------------------
    conn.execute("""
    CREATE TABLE IF NOT EXISTS announcements (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title TEXT NOT NULL,
        body TEXT NOT NULL,
        image_path TEXT,                      -- opsiyonel resim yolu
        is_active INTEGER NOT NULL DEFAULT 1, -- aktif/pasif
        created_at TEXT NOT NULL,
        created_by INTEGER,
        updated_at TEXT,
        updated_by INTEGER,
        FOREIGN KEY(created_by) REFERENCES users(id),
        FOREIGN KEY(updated_by) REFERENCES users(id)
    )
    """)

    # -----------------------------
    # Reminders (Süre Takibi / Hatırlatma)
    # -----------------------------
    conn.execute("""
    CREATE TABLE IF NOT EXISTS reminders (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title TEXT NOT NULL,                 -- Örn: "34ABC123 - Seyrüsefer"
        due_date TEXT NOT NULL,              -- "YYYY-MM-DD"
        remind_days TEXT NOT NULL DEFAULT '30,7,1,0',  -- kaç gün kala mail gitsin
        target_roles TEXT NOT NULL DEFAULT 'owner',    -- örn: "owner,accounting"
        target_emails TEXT,                  -- opsiyon: "a@x.com,b@y.com"
        is_active INTEGER NOT NULL DEFAULT 1,
        created_at TEXT NOT NULL
    )
    """)

    conn.execute("""
    CREATE TABLE IF NOT EXISTS reminder_notifications (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        reminder_id INTEGER NOT NULL,
        notify_date TEXT NOT NULL,           -- "YYYY-MM-DD" (due_date - gün)
        sent_at TEXT,                        -- gönderildiyse timestamp
        UNIQUE(reminder_id, notify_date)
    )
    """)

    # Reminders: sahiplik + hedef kullanıcılar (yetkilendirme için)
    try:
        conn.execute("ALTER TABLE reminders ADD COLUMN created_by INTEGER")
    except Exception:
        pass

    try:
        conn.execute("ALTER TABLE reminders ADD COLUMN scope TEXT NOT NULL DEFAULT 'private'")
    except Exception:
        pass

    conn.execute("""
    CREATE TABLE IF NOT EXISTS reminder_targets (
        reminder_id INTEGER NOT NULL,
        user_id INTEGER NOT NULL,
        UNIQUE(reminder_id, user_id)
    )
    """)

    exec_sql("""
    CREATE TABLE IF NOT EXISTS dino_scores (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL,
        score INTEGER NOT NULL,
        created_at TEXT NOT NULL DEFAULT (datetime('now'))
    )
    """)
    exec_sql("CREATE INDEX IF NOT EXISTS idx_dino_scores_user_id ON dino_scores(user_id)")
    exec_sql("CREATE INDEX IF NOT EXISTS idx_dino_scores_score ON dino_scores(score)")    

    conn.commit()
    conn.close()

def ensure_default_owner():
    row = query_one("SELECT COUNT(*) as c FROM users")
    if row and row["c"] == 0:
        owner_user = os.environ.get("IK_OWNER_USER", "admin")
        owner_pass = os.environ.get("IK_OWNER_PASS", "Admin123+")
        owner_name = os.environ.get("IK_OWNER_NAME", "Patron")
        owner_mail = os.environ.get("IK_OWNER_EMAIL", "")

        exec_sql("""
        INSERT INTO users (full_name, username, password_hash, role, is_active, email, annual_leave_days, can_qr, qr_secret)
        VALUES (?, ?, ?, ?, 1, ?, 15, 1, ?)
        """, (
            owner_name,
            owner_user.lower().strip(),
            generate_password_hash(owner_pass),
            ROLE_OWNER,
            owner_mail,
            secrets.token_hex(8)
        ))


# -----------------------------
# Auth & RBAC
# -----------------------------
def current_user():
    uid = session.get("user_id") or session.get("uid")
    if not uid:
        return None
    return query_one("SELECT * FROM users WHERE id=? AND is_active=1", (uid,))

def is_admin_user():
    role = session.get("role")
    return role in (ROLE_OWNER, ROLE_ACCOUNTING, ROLE_MANAGER)

def login_required(fn):
    def wrapper(*args, **kwargs):
        if not current_user():
            return redirect(url_for("login", next=request.path))
        return fn(*args, **kwargs)
    wrapper.__name__ = fn.__name__
    return wrapper

def role_required(*roles):
    def deco(fn):
        def wrapper(*args, **kwargs):
            u = current_user()
            if not u:
                return redirect(url_for("login", next=request.path))
            if u["role"] not in roles:
                abort(403)
            return fn(*args, **kwargs)
        wrapper.__name__ = fn.__name__
        return wrapper
    return deco


# -----------------------------
# Mail
# -----------------------------
def send_mail(to_email: str, subject: str, body_html: str):
    if not SMTP_HOST or not SMTP_USER or not SMTP_PASS:
        return

    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"] = MAIL_FROM
        msg["To"] = to_email

        msg.attach(MIMEText(body_html, "html", "utf-8"))

        with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=15) as s:
            s.ehlo()
            s.starttls()
            s.ehlo()
            s.login(SMTP_USER, SMTP_PASS)
            s.sendmail(MAIL_FROM, [to_email], msg.as_string())

    except Exception as e:
        # Mail atılamadı diye izin talebi/işlem çökmesin
        try:
            app.logger.exception("Mail gönderilemedi: %s", e)
        except Exception:
            print("Mail gönderilemedi:", e)
        return


def send_mail_many(to_addrs, subject: str, body: str):
    if not to_addrs:
        return
    for a in to_addrs:
        try:
            send_mail(a, subject, body)
        except Exception:
            pass

def get_owner_email():
    r = query_one("SELECT email FROM users WHERE role=? AND is_active=1 ORDER BY id ASC LIMIT 1", (ROLE_OWNER,))
    return (r["email"] if r else "") or ""

def get_accounting_emails():
    rows = query_all("SELECT email FROM users WHERE role=? AND is_active=1 AND email IS NOT NULL AND email!=''", (ROLE_ACCOUNTING,))
    return [r["email"] for r in rows]
def _parse_int_list(csv_text: str):
    out = []
    for x in (csv_text or "").split(","):
        x = x.strip()
        if not x:
            continue
        try:
            out.append(int(x))
        except Exception:
            pass
    return out


def _parse_str_list(csv_text: str):
    out = []
    for x in (csv_text or "").split(","):
        x = x.strip()
        if x:
            out.append(x)
    return out


def _get_user_emails_by_roles(roles_csv: str):
    roles = [r.strip() for r in (roles_csv or "").split(",") if r.strip()]
    if not roles:
        return []
    qmarks = ",".join(["?"] * len(roles))
    rows = query_all(f"""
        SELECT email FROM users
        WHERE is_active=1
          AND email IS NOT NULL AND TRIM(email) <> ''
          AND role IN ({qmarks})
    """, tuple(roles))
    out = []
    for r in rows:
        e = (r["email"] or "").strip()
        if e:
            out.append(e)
    return out


def _get_user_email_by_id(uid: int):
    row = query_one("""
        SELECT email FROM users
        WHERE id=? AND is_active=1
          AND email IS NOT NULL AND TRIM(email) <> ''
    """, (uid,))
    if not row:
        return ""
    # sqlite3.Row -> dict gibi davranır ama .get yok
    email = (row["email"] or "").strip()
    return email


def _get_target_emails_for_reminder(rem: dict):
    """
    Basit hedefleme:
    - private: sadece created_by (kendi maili) + target_emails (varsa)
    - team/global: target_roles'daki rollerin mailleri + target_emails
    Not: team dağıtımı (manager/responsible üzerinden reminder_targets) sonraki adımda.
    """
    scope = (rem.get("scope") or "private").strip()
    created_by = rem.get("created_by")
    target_roles = (rem.get("target_roles") or "").strip()
    extra = _parse_str_list(rem.get("target_emails") or "")

    emails = set()

    if scope == "private":
        if created_by:
            e = _get_user_email_by_id(int(created_by))
            if e:
                emails.add(e)
        for e in extra:
            emails.add(e)
        return sorted(emails)

    # team/global -> roller + extra
    for e in _get_user_emails_by_roles(target_roles):
        emails.add(e)
    for e in extra:
        emails.add(e)
    return sorted(emails)


def run_reminders_mail_worker(today_str: str = None):
    """
    Bugün gönderilmesi gereken hatırlatma maillerini gönderir.
    - remind_days içindeki günlere göre notify_date hesaplar (due_date - gün)
    - reminder_notifications tablosu ile aynı notify_date için tekrar mail atmayı engeller
    """
    # bugün
    today = date.today()
    if today_str:
        # YYYY-MM-DD
        try:
            today = datetime.strptime(today_str, "%Y-%m-%d").date()
        except Exception:
            pass
    today_s = today.strftime("%Y-%m-%d")

    # aktif reminderlar
    rows = query_all("""
        SELECT id, title, due_date, remind_days, target_roles, target_emails,
               is_active, created_by, scope
        FROM reminders
        WHERE is_active=1
    """)

    sent_count = 0

    for rem in rows:
        rem = dict(rem)
        rid = int(rem["id"])
        due_s = (rem.get("due_date") or "").strip()
        if not due_s:
            continue

        try:
            due_d = datetime.strptime(due_s, "%Y-%m-%d").date()
        except Exception:
            continue

        days_list = _parse_int_list(rem.get("remind_days") or "30,7,1,0")
        if not days_list:
            days_list = [0]

        # bugüne denk geliyor mu?
        should_send_today = False
        for d in days_list:
            nd = (due_d - timedelta(days=int(d))).strftime("%Y-%m-%d")
            if nd == today_s:
                should_send_today = True
                break
        if not should_send_today:
            continue

        # daha önce gönderildi mi?
        already = query_one("""
            SELECT id, sent_at FROM reminder_notifications
            WHERE reminder_id=? AND notify_date=?
        """, (rid, today_s))
        if already and already["sent_at"]:
            continue

        to_addrs = _get_target_emails_for_reminder(rem)
        if not to_addrs:
            # hedef yoksa yine de "gönderildi" saymayalım
            continue

        subj = f"[IK][REMINDER] {rem.get('title') or 'Hatırlatma'}"
        body = f"""
        <div style="font-family:Arial,Helvetica,sans-serif;line-height:1.5">
          <h3 style="margin:0 0 10px 0">Hatırlatma</h3>
          <p><b>Başlık:</b> {html_escape(rem.get('title') or '')}</p>
          <p><b>Son Tarih:</b> {html_escape(due_s)}</p>
          <p style="color:#6b7280;font-size:13px;margin-top:14px">
            Bu mail İK Portal Hatırlatmalar modülü tarafından otomatik gönderilmiştir.
          </p>
        </div>
        """

        # mail gönder (tek tek; daha garantili)
        ok_cnt = 0
        for _to in to_addrs:
            try:
                send_mail(_to, subj, body)
                ok_cnt += 1
            except Exception as e:
                print("REMINDER_SEND_ONE_ERROR to:", _to, "err:", repr(e), flush=True)

        print("REMINDER_SEND to:", ",".join(to_addrs), "subj:", subj, "sent:", ok_cnt, flush=True)

        # gönderim kaydı (insert or ignore + update)
        exec_sql("""
            INSERT OR IGNORE INTO reminder_notifications (reminder_id, notify_date, sent_at)
            VALUES (?, ?, ?)
        """, (rid, today_s, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        exec_sql("""
            UPDATE reminder_notifications
            SET sent_at=COALESCE(sent_at, ?)
            WHERE reminder_id=? AND notify_date=?
        """, (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), rid, today_s))

        sent_count += 1

    return sent_count

def get_it_emails():
    rows = query_all(
        "SELECT email FROM users WHERE role=? AND is_active=1 AND email IS NOT NULL AND email!=''",
        (ROLE_IT,)
    )
    return [r["email"] for r in rows]


def get_user_email(user_id: int):
    r = query_one("SELECT email FROM users WHERE id=? AND is_active=1", (user_id,))
    return (r["email"] if r else "") or ""


# -----------------------------
# UI (light corporate)
# -----------------------------
def html_escape(s: str) -> str:
    return (s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

# ✅ Mail helpers (ŞIK / ALT ALTA)
def nl2br(s: str) -> str:
    return html_escape(s).replace("\n", "<br>")

def ik_mail_template(title: str, intro: str, rows: list, stage_text: str = "", button_text: str = "", button_url: str = "", footer: str = "İK Portal"):
    """
    rows: [("Etiket", "Değer (HTML olabilir)"), ...]
    Not: rows içindeki value tarafı HTML olabilir (nl2br vb). Key tarafı escape edilir.
    """
    tr_html = ""
    for k, v in rows:
        tr_html += f"""
        <tr>
          <td style="padding:7px 0; color:#64748b; width:140px; vertical-align:top;">{html_escape(str(k))}</td>
          <td style="padding:7px 0; font-weight:700;">{v}</td>
        </tr>
        """

    stage_block = ""
    if stage_text:
        stage_block = f"""
        <div style="margin-top:14px; border:1px solid #e5e7eb; border-radius:12px; padding:12px 14px; background:#ffffff;">
          <div style="font-weight:800; margin-bottom:8px;">Onay Aşaması</div>
          <div style="margin-bottom:10px;">
            <span style="display:inline-block; padding:6px 12px; border-radius:999px; background:#fff7ed; border:1px solid #fed7aa; color:#9a3412; font-weight:800;">
              {html_escape(stage_text)}
            </span>
          </div>
          {"<a href='"+html_escape(button_url)+"' style='display:inline-block; padding:10px 14px; border-radius:12px; background:#1d4ed8; color:#ffffff; font-weight:800; text-decoration:none;'>"+html_escape(button_text)+"</a>" if (button_text and button_url) else ""}
          {("<div style='margin-top:10px; font-size:12px; color:#64748b;'>Buton çalışmazsa link: "+html_escape(button_url)+"</div>") if (button_text and button_url) else ""}
        </div>
        """

    return f"""
<div style="font-family:Segoe UI,Roboto,Arial; font-size:14px; color:#0f172a; background:#f6f8fc; padding:16px 0;">
  <div style="max-width:680px; margin:0 auto; border:1px solid #e5e7eb; border-radius:14px; overflow:hidden; background:#ffffff;">
    <div style="background:#1d4ed8; color:white; padding:16px 18px;">
      <div style="font-size:16px; font-weight:900; letter-spacing:.2px;">{html_escape(APP_TITLE)}</div>
      <div style="opacity:.92; margin-top:4px;">{html_escape(title)}</div>
    </div>

    <div style="padding:18px;">
      <p style="margin:0 0 10px 0;">Merhaba,</p>
      <p style="margin:0 0 14px 0; color:#334155;">{html_escape(intro)}</p>

      <div style="border:1px solid #e5e7eb; border-radius:12px; padding:12px 14px; background:#f8fafc;">
        <div style="font-weight:900; margin-bottom:8px;">Talep Bilgileri</div>
        <table style="width:100%; border-collapse:collapse; font-size:14px;">
          {tr_html}
        </table>
      </div>

      {stage_block}

      <div style="margin-top:16px; color:#94a3b8; font-size:12px;">
        {html_escape(footer)}
      </div>
    </div>
  </div>
</div>
"""

def format_status_pill(status: str) -> str:
    s = status or ""
    cls = "warn"
    if s == STATUS_APPROVED or s == ADV_STATUS_APPROVED:
        cls = "ok"
    elif s == STATUS_REJECTED or s == ADV_STATUS_REJECTED:
        cls = "bad"
    return f"<span class='pill {cls}'>{html_escape(s)}</span>"

def get_active_announcements():
    return query_all("""
        SELECT id, title, body, image_path
        FROM announcements
        WHERE is_active = 1
        ORDER BY id DESC
    """)

def render_page(title, body_html, user=None):
    u = user or current_user()
    nav = ""
    if u:
        links_top = []
        links_groups = []

        def add_top(href, text):
            links_top.append((href, text))

        def add_group(title, items):
            # items: [(href, text), ...]
            if items:
                links_groups.append((title, items))

        # Herkes
        add_top("/", "Ana Sayfa")

        # ✅ İZİNLER
        izin_items = []
        if u["role"] != ROLE_OWNER:
            izin_items.append(("/leave/my", "İzinlerim"))
            izin_items.append(("/leave/new", "İzin Talep Et"))

        if u["role"] in (ROLE_OWNER, ROLE_ACCOUNTING, ROLE_MANAGER, ROLE_RESPONSIBLE):
            izin_items.append(("/leave/admin", "İzin Yönetimi"))

        if u["role"] in (ROLE_OWNER, ROLE_ACCOUNTING, ROLE_MANAGER):
            izin_items.append(("/leave/calendar", "İzin Takvimi"))

        add_group("İzinler", izin_items)

        # ✅ MESAİ
        mesai_items = []
        if u["can_qr"] == 1:
            mesai_items.append(("/qr", "QR ile Mesai"))

        if u["role"] in (ROLE_OWNER, ROLE_ACCOUNTING):
            mesai_items.append(("/attendance", "Mesai Raporu"))

        if u["role"] == ROLE_RESPONSIBLE:
            mesai_items.append(("/attendance/team", "Mesai (Ekibim)"))

        add_group("Mesai", mesai_items)

        # ✅ AVANS
        avans_items = []
        if u["role"] != ROLE_OWNER:
            avans_items.append(("/advance/my", "Avanslarım"))
            avans_items.append(("/advance/new", "Avans Talep Et"))

        if u["role"] == ROLE_ACCOUNTING:
            avans_items.append(("/advance/accounting", "Avans Yönetimi"))

        if u["role"] == ROLE_OWNER:
            avans_items.append(("/advance/owner", "Avans Onay"))

        add_group("Avans", avans_items)

        # ✅ IT
        it_items = [
            ("/it/my", "IT Taleplerim"),
            ("/it/new", "IT Talep Aç"),
        ]
        if u["role"] == ROLE_IT:
            it_items.append(("/it/admin", "IT Talep Yönetimi"))

        add_group("IT", it_items)

        # 🔔 Hatırlatmalar
        add_top("/reminders", "Hatırlatmalar")

        # 📣 Duyurular (Owner + Accounting)
        if u["role"] in (ROLE_OWNER, ROLE_ACCOUNTING):
            add_top("/admin/announcements", "Duyurular")

        # Kullanıcılar
        if u["role"] in (ROLE_OWNER, ROLE_ACCOUNTING):
            add_top("/users", "Kullanıcılar")

        # HTML üret
        nav_html = ""
        for href, text in links_top:
            nav_html += f"<a class='navlink' href='{href}'>{html_escape(text)}</a>"

        for gtitle, items in links_groups:
            dd_links = "".join([
                f"<a href='{href}'>{html_escape(text)}</a>"
                for href, text in items
            ])
            nav_html += f"""
            <div class="navGroup">
              <button class="navDropBtn" type="button">{html_escape(gtitle)} ▾</button>
              <div class="navDrop">
                {dd_links}
              </div>
            </div>
            """

        nav_links = nav_html

        # ✅ Responsive hamburger menü (mobil)
        nav = f"""
        <div class="topbar">
          <div class="brand">{APP_TITLE}</div>

          <input id="navToggle" class="navToggle" type="checkbox">
          <label class="hamburger" for="navToggle" aria-label="Menüyü Aç/Kapat">
            <span></span><span></span><span></span>
          </label>

          <div class="nav" id="topNav">{nav_links}</div>

          <div class="userbox">
            <span class="muted userName">{html_escape(u['full_name'])}</span>
            <a class="navlink" href="/logout">Çıkış</a>
          </div>
        </div>
        """
    else:
        nav = f"""
        <div class="topbar">
          <div class="brand">{APP_TITLE}</div>
          <div class="nav"></div>
          <div class="userbox"></div>
        </div>
        """

    return f"""
    <!doctype html>
    <html lang="tr">
    <head>
      <meta charset="utf-8"/>
      <meta name="viewport" content="width=device-width, initial-scale=1"/>
      <title>{html_escape(title)} - {APP_TITLE}</title>
      <style>
        :root {{
          --bg:#f6f8fc;
          --card:#ffffff;
          --text:#0f172a;
          --muted:#64748b;
          --line:rgba(15,23,42,.10);
          --primary:#1d4ed8;
          --primary2:#1e40af;
          --ok:#16a34a;
          --bad:#dc2626;
          --warn:#f59e0b;
          --shadow: 0 10px 28px rgba(2,6,23,.08);
        }}
        *{{box-sizing:border-box}}
        body{{
            margin:0;
            background:
              radial-gradient(1200px 700px at 20% 10%, rgba(14,165,233,.15), transparent 60%),
              radial-gradient(1000px 600px at 80% 20%, rgba(59,130,246,.12), transparent 55%),
              linear-gradient(180deg,#0f172a 0%, #0b1220 100%);
            color:#f1f5f9;
            font-family:system-ui,-apple-system,Segoe UI,Roboto,Arial;
        }}
        a{{color:var(--primary);text-decoration:none}}

        .topbar{{
            display:flex;
            align-items:center;
            justify-content:space-between;
            padding:14px 18px;
            border-bottom:1px solid rgba(148,163,184,.15);
            backdrop-filter: blur(12px);
            background:rgba(15,23,42,.85);
            box-shadow:0 8px 30px rgba(0,0,0,.45);
            position:sticky;
            top:0;
            z-index:30;
        }}

        .brand{{font-weight:800;letter-spacing:.2px; white-space:nowrap}}
        .nav{{
            display:flex;
            gap:10px;
            flex-wrap:nowrap;
            align-items:center;
        }}

        .topbar .nav > a{{
            display:inline-flex;
            align-items:center;
            justify-content:center;
            height:44px;
            padding:0 14px;
            border:1px solid #dbe3ff;
            border-radius:18px;
            background:#f4f7ff;
            white-space:nowrap;
            color:#1e293b;
            font-weight:600;
            line-height:1;
            vertical-align:middle;
            transition:all .15s ease;
        }}
        .topbar .nav > a:hover{{
            background:#e6ecff;
            border-color:#c7d2fe;
        }}

        .navGroup{{
            position:relative;
            display:inline-flex;
            align-items:center;
        }}

        .navDropBtn{{
            display:inline-flex;
            align-items:center;
            justify-content:center;
            height:44px;
            padding:0 14px;
            border:1px solid #dbe3ff;
            border-radius:18px;
            background:#f4f7ff;
            white-space:nowrap;
            cursor:pointer;
            font:inherit;
            font-weight:600;
            color:#1e293b;
            line-height:1;
            vertical-align:middle;
            transition:all .15s ease;
            appearance:none;
        }}
        .navDropBtn:hover{{
            background:#e6ecff;
            border-color:#c7d2fe;
        }}

        .nav a.btn2{{
            padding:8px 10px;
            border:1px solid var(--line);
            border-radius:12px;
            background:#fff;
            white-space:nowrap;
            color:var(--primary);
        }}
        .nav a.btn2:hover{{
            background:#f3f6ff;
        }}

        .navDropBtn:hover{{background:#f3f6ff}}

        .navDrop{{
          display:none;
          position:absolute;
          top:100%;
          left:0;
          min-width:220px;
          background:#ffffff;
          border:1px solid rgba(15,23,42,.10);
          border-radius:16px;
          box-shadow:0 18px 40px rgba(2,6,23,.12);
          padding:10px;
          z-index:999;
        }}

        .navGroup:hover .navDrop{{display:block}}

        .navDrop a{{
          display:block;
          padding:10px 10px;
          border-radius:12px;
          color:var(--text);
          border:1px solid transparent;
          white-space:nowrap;
        }}
        .navDrop a:hover{{
          background:#f3f6ff;
          border-color:var(--line);
        }}

          /* Mobil görünüm */
          @media (max-width:900px){{
            .topbar{{
              display:flex;
              flex-direction:column;
              align-items:stretch;
              gap:10px;
              padding:12px 12px;
              width:100%;
              box-sizing:border-box;
              overflow:visible;
            }}

            .brand{{
              font-size:16px;
            }}

            /* Menü yatay kaydırmalı olsun */
            .nav{{
              display:flex;
              flex-wrap:wrap;
              justify-content:flex-start;
              align-items:center;
              gap:10px;
              white-space:normal;
              overflow:visible;
              width:100%;
              box-sizing:border-box;
            }}
            .userbox{{
              width:100%;
              display:flex;
              justify-content:space-between;
              align-items:center;
              flex-wrap:wrap;
              gap:8px;
            }}
            .userbox a{{white-space:nowrap;}}

            .navGroup{{
              flex:0 0 auto;
            }}

            .nav a{{
              flex:0 0 auto;
            }}

            /* Dropdown mobilde sabit panel gibi açılsın */
            .navDrop{{
              position:fixed;
              left:12px;
              right:12px;
              top:70px;
              min-width:unset;
              max-height:60vh;
              overflow:auto;
              z-index:9999;
            }}
          }}

        .userbox{{display:flex;align-items:center;gap:10px; flex-wrap:wrap; justify-content:flex-end}}
        .container{{max-width:1100px;margin:22px auto;padding:0 14px}}
        .pageHeader{{
            padding:28px 0 12px;
        }}
        .pageHeaderInner{{
            max-width:1100px;
            margin:0 auto;
            padding:0 18px;
        }}
        .pageTitle{{
            font-size:26px;
            font-weight:800;
            letter-spacing:-0.3px;
        }}
        .card{{
            background:#111827;
            border:1px solid rgba(148,163,184,.15);
            border-radius:18px;
            padding:18px;
            box-shadow:0 20px 45px rgba(0,0,0,.45);
        }}
        .grid{{display:grid;grid-template-columns:repeat(12,1fr);gap:14px}}
        .two{{display:grid;grid-template-columns:1fr 1fr;gap:14px}}
        .h1{{font-size:20px;font-weight:800;margin:0 0 10px}}
        .muted{{color:#cbd5e1}}
        input,select,textarea{{width:100%;padding:10px 12px;border-radius:12px;border:1px solid var(--line);background:#fff;color:var(--text);outline:none}}
        textarea{{min-height:90px;resize:vertical}}
        .row{{display:grid;grid-template-columns:1fr 1fr;gap:12px}}
        .btn{{display:inline-block;padding:10px 14px;border-radius:12px;border:1px solid rgba(29,78,216,.20);background:linear-gradient(180deg,var(--primary),var(--primary2));color:white;font-weight:700;cursor:pointer}}
        .btn:hover{{filter:brightness(1.05)}}
        .btn2{{display:inline-block;padding:10px 14px;border-radius:12px;border:1px solid var(--line);background:#fff;color:var(--text);font-weight:700}}
        .btn2:hover{{background:#f3f6ff}}
        .btn-excel{{
          background:#1D6F42;
          border:1px solid #1D6F42;
          color:#ffffff;
        }}
        .btn-excel:hover{{
          background:#155732;
          border-color:#155732;
          color:#ffffff;
        }}

        /* ✅ Table mobilde taşmasın: yatay kaydır */
        table{{width:100%;border-collapse:collapse;display:block;overflow-x:auto;max-width:100%;-webkit-overflow-scrolling:touch}}
        th,td{{border-bottom:1px solid var(--line);padding:10px 8px;text-align:left;font-size:14px;white-space:nowrap}}
        table td{{
            color:#f8fafc;
        }}

        table th{{
            color:#e2e8f0;
        }}
        .pill{{
            display:inline-block;
            padding:4px 10px;
            border-radius:999px;
            border:1px solid rgba(148,163,184,.25);
            background:rgba(59,130,246,.12);
            font-size:12px;
            color:#e2e8f0;
        }}
        .pill.ok{{border-color:rgba(22,163,74,.25);background:rgba(22,163,74,.08);color:var(--ok)}}
        .pill.bad{{border-color:rgba(220,38,38,.25);background:rgba(220,38,38,.08);color:var(--bad)}}
        .pill.warn{{border-color:rgba(245,158,11,.25);background:rgba(245,158,11,.10);color:#92400e}}
        .note{{
            border:1px dashed rgba(148,163,184,.25);
            background:rgba(30,41,59,.55);
            padding:12px;
            border-radius:14px;
            color:#e2e8f0;
        }}

        /* ✅ Hamburger (mobil menü) */
        .navToggle{{display:none}}
        .hamburger{{
          display:none;
          width:42px; height:38px;
          border:1px solid var(--line);
          border-radius:12px;
          background:#fff;
          align-items:center;
          justify-content:center;
          gap:5px;
          cursor:pointer;
        }}
        .hamburger span{{
          display:block;
          width:18px;
          height:2px;
          background:rgba(15,23,42,.70);
          border-radius:2px;
        }}

        @media (max-width:900px){{
          .row{{grid-template-columns:1fr}}
          .two{{grid-template-columns:1fr}}

          /* menü mobilde dropdown */
          .hamburger{{display:flex}}
          .nav{{
            width:100%;
            display:none;
            padding:10px 0 2px;
            border-top:1px solid var(--line);
          }}
          .navlink{{padding:10px 12px}}
          .navToggle:checked ~ .nav{{display:flex}}
          .userName{{display:none}} /* çok dar ekranda isim taşmasın */
        }}

        @media (max-width:520px){{
          .container{{margin:16px auto}}
          .topbar{{padding:12px 12px}}
          .brand{{font-size:14px}}
          .btn,.btn2{{width:100%; text-align:center}}
        }}
      </style>
    </head>
    <body>
      {nav}
      <div class="pageHeader">
        <div class="pageHeaderInner">
          <h1 class="pageTitle">{title}</h1>
        </div>
      </div>
      <div class="container">
        {body_html}
      </div>
    </body>
    </html>
    """

def device_fingerprint():
    ip = request.headers.get("X-Forwarded-For", request.remote_addr) or ""
    if "," in ip:
        ip = ip.split(",")[0].strip()
    ua = request.headers.get("User-Agent", "")[:180]
    return ip.strip(), ua.strip()


# -----------------------------
# Annual leave calculations
# -----------------------------
def _full_years_between(start_d: date, end_d: date) -> int:
    """Tam yıl hesabı (doğum günü mantığı): yıl farkı, ay/gün gelmediyse 1 düş."""
    y = end_d.year - start_d.year
    if (end_d.month, end_d.day) < (start_d.month, start_d.day):
        y -= 1
    return max(0, y)

def annual_entitlement_days(user_row, on_date: date = None) -> float:
    """
    Otomatik yıllık izin hakkı (5 yılda +3 gün kuralı):
      - 0-4 yıl: 15
      - 5-9 yıl: 18
      - 10-15 yıl: 21
      - 15-19 yıl: 24
      - 20-24 yıl: 27
      ... devam eder

    Manuel override (annual_leave_override) doluysa otomatiği ezer.
    hire_date boş/bozuksa fallback: annual_leave_days
    """
    if on_date is None:
        on_date = date.today()

    # ✅ 1) Manuel override varsa onu kullan
    try:
        if user_row and ("annual_leave_override" in user_row.keys()):
            ov = user_row["annual_leave_override"]
            if ov is not None and str(ov).strip() != "":
                ov_i = int(ov)
                if ov_i >= 0:
                    return float(ov_i)
    except Exception:
        pass

    # ✅ 2) hire_date varsa otomatik hesap
    hire_s = ""
    try:
        hire_s = (user_row["hire_date"] if user_row and ("hire_date" in user_row.keys()) else "") or ""
        hire_s = hire_s.strip()
    except Exception:
        hire_s = ""

    if hire_s:
        try:
            hd = datetime.strptime(hire_s, "%Y-%m-%d").date()
            years = _full_years_between(hd, on_date)

            # 0-4 -> 15, 5-9 -> 18 ... (her 5 yılda +3)
            base = 15
            add_blocks = years // 5  # 0..4->0, 5..9->1, 10..14->2 ...
            return float(base + (add_blocks * 3))
        except Exception:
            pass

    # ✅ 3) hire_date yoksa/bozuksa manuel annual_leave_days'e fallback
    try:
        return float(int(user_row["annual_leave_days"] or 15))
    except Exception:
        return 15.0


def days_between_inclusive(start_date: str, end_date: str) -> int:
    d1 = datetime.strptime(start_date, "%Y-%m-%d").date()
    d2 = datetime.strptime(end_date, "%Y-%m-%d").date()
    if d2 < d1:
        return 0
    return (d2 - d1).days + 1

def used_annual_leave_days(user_id: int, year: int) -> float:
    y1 = f"{year}-01-01"
    y2 = f"{year}-12-31"
    rows = query_all("""
        SELECT start_date, end_date, is_half_day
        FROM leave_requests
        WHERE user_id=?
          AND status=?
          AND leave_type='Yillik'
    """, (user_id, STATUS_APPROVED))

    total = 0.0
    for r in rows:
        if r["end_date"] < y1 or r["start_date"] > y2:
            continue

        # ✅ Yarım gün
        if int(r["is_half_day"] or 0) == 1:
            total += 0.5
            continue

        # ✅ Tam gün (eski hesap)
        s = max(r["start_date"], y1)
        e = min(r["end_date"], y2)
        total += float(days_between_inclusive(s, e))

    return total

def remaining_annual_leave_days(user_row, year: int) -> float:
    # ✅ HAK: hire_date'e göre otomatik
    annual = float(annual_entitlement_days(user_row, date(year, 12, 31)))
    used = float(used_annual_leave_days(int(user_row["id"]), year))
    rem = annual - used
    return rem if rem > 0 else 0.0


# -----------------------------
# QR token logic
# -----------------------------
def make_token(qr_secret: str, d: date) -> str:
    raw = f"{qr_secret}_{d.isoformat()}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()

def find_user_by_token(token: str, d: date):
    users = query_all("""
        SELECT id, full_name, qr_secret
        FROM users
        WHERE is_active=1 AND can_qr=1 AND qr_secret IS NOT NULL AND qr_secret!=''
    """)
    for u in users:
        expected = make_token(u["qr_secret"], d)
        if secrets.compare_digest(expected, token):
            return u
    return None


# -----------------------------
# Attendance logic
# -----------------------------
def next_attendance_tur(user_id: int, tarih: str):
    rows = query_all("SELECT tur FROM attendance WHERE user_id=? AND tarih=? ORDER BY id ASC", (user_id, tarih))
    count = len(rows)
    if count >= len(TUR_SEQUENCE):
        return None
    return TUR_SEQUENCE[count]

def device_used_by_another_user_today(tarih: str, ip: str, cihaz: str, user_id: int):
    row = query_one("""
    SELECT a.user_id, u.full_name
    FROM attendance a
    JOIN users u ON u.id=a.user_id
    WHERE a.tarih=? AND a.ip=? AND a.cihaz=? AND a.user_id != ?
    LIMIT 1
    """, (tarih, ip, cihaz, user_id))
    return row

def insert_attendance(user_id: int, tur: str, ip: str, cihaz: str):
    now = datetime.now()
    tarih = now.strftime("%Y-%m-%d")
    saat = now.strftime("%H:%M")
    exec_sql("""
    INSERT INTO attendance (user_id, tarih, saat, tur, ip, cihaz, created_at)
    VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (user_id, tarih, saat, tur, ip, cihaz, now.isoformat(timespec="seconds")))
    return tarih, saat

def last_attendance_dt(user_id: int):
    row = query_one("""
        SELECT created_at
        FROM attendance
        WHERE user_id=?
        ORDER BY id DESC
        LIMIT 1
    """, (user_id,))
    if not row or not row["created_at"]:
        return None
    try:
        return datetime.fromisoformat(row["created_at"])
    except Exception:
        return None


# -----------------------------
# Helpers for multi-step leave approval (✅ FIX)
# -----------------------------
def can_act_on_leave(decider_row, leave_row) -> bool:
    """Only the current stage owner can act (owner/accounting for OWNER stage)."""
    if leave_row["status"] != STATUS_PENDING:
        return False

    stage = (leave_row["stage"] or "").strip() or STAGE_OWNER
    pending_with = leave_row["pending_with"]

    if stage == STAGE_RESPONSIBLE:
        return pending_with == decider_row["id"]
    if stage == STAGE_MANAGER:
        return pending_with == decider_row["id"]
    # OWNER stage:
    if stage == STAGE_OWNER:
        return decider_row["role"] in (ROLE_OWNER, ROLE_ACCOUNTING)
    return False

def first_leave_stage_for_requester(requester_row):
    """
    Personel -> varsa sorumlu, yoksa müdür, yoksa owner
    Diğer roller -> owner
    """
    if requester_row["role"] in (ROLE_PERSONNEL, ROLE_IT):
        rid = requester_row["responsible_id"]
        if rid:
            return (STAGE_RESPONSIBLE, int(rid))
        mid = requester_row["manager_id"]
        if mid:
            return (STAGE_MANAGER, int(mid))
        return (STAGE_OWNER, None)
    return (STAGE_OWNER, None)

def next_leave_stage_after_responsible(requester_row):
    """
    Sorumlu onayından sonra:
    - Müdür varsa müdüre
    - Yoksa owner'a
    """
    mid = requester_row["manager_id"]
    if mid:
        return (STAGE_MANAGER, int(mid))
    return (STAGE_OWNER, None)

def stage_label(stage: str) -> str:
    s = (stage or "").strip() or STAGE_OWNER
    if s == STAGE_RESPONSIBLE:
        return "Sorumlu Onayı"
    if s == STAGE_MANAGER:
        return "Müdür Onayı"
    if s == STAGE_OWNER:
        return "Patron/Muhasebe Onayı"
    return s


# -----------------------------
# Routes
# -----------------------------
@app.route("/health")
def health():
    return "OK"

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = (request.form.get("username") or "").strip().lower()
        password = (request.form.get("password") or "").strip()
        u = query_one("SELECT * FROM users WHERE username=? AND is_active=1", (username,))
        if u and check_password_hash(u["password_hash"], password):
            session["uid"] = u["id"]
            session["user_id"] = u["id"]   # ✅ uyumluluk için (duyurular vs.)
            # role session'a yaz (sqlite Row'da .get yok)
            session["role"] = u["role"] if ("role" in u.keys()) else None
            nxt = request.args.get("next") or "/"
            return redirect(nxt)
        body = """
        <div class="card" style="max-width:520px;margin:0 auto;">
          <h1 class="h1">Giriş</h1>
          <div class="pill bad">Hatalı kullanıcı adı veya şifre</div>
          <div style="margin-top:12px"><a class="btn2" href="/login">Tekrar Dene</a></div>
        </div>
        """
        return render_page("Giriş", body, user=None)

    body = """
    <div class="card" style="max-width:520px;margin:0 auto;">
      <h1 class="h1">İK Portal Giriş</h1>
      <p class="muted">Kullanıcı adınız ve şifrenizle giriş yapın.</p>
      <form method="post">
        <div class="row">
          <div>
            <label class="muted">Kullanıcı Adı</label>
            <input name="username" autocomplete="username" required>
          </div>
          <div>
            <label class="muted">Şifre</label>
            <input name="password" type="password" autocomplete="current-password" required>
          </div>
        </div>
        <div style="margin-top:12px">
          <button class="btn" type="submit">Giriş Yap</button>
        </div>
      </form>
      <div class="note" style="margin-top:14px">
        <div class="muted">Sorun yaşarsanız yetkili birimle iletişime geçin.</div>
      </div>
    </div>
    """
    return render_page("Giriş", body, user=None)

# -----------------------------
# ✅ IT TALEP (NEW)
# -----------------------------
IT_PRIORITIES = ["Düşük", "Orta", "Yüksek", "Acil"]
IT_STATUSES = ["Açık", "Beklemede", "Yönlendirildi", "Kapalı"]


def _can_view_ticket(u, ticket_row) -> bool:
    if not u or not ticket_row:
        return False
    if u["role"] == ROLE_IT:
        return True
    return int(ticket_row["user_id"]) == int(u["id"])


@app.route("/it/new", methods=["GET", "POST"])
@login_required
def it_new():
    u = current_user()

    # GET -> form
    if request.method != "POST":
        body = f"""
        <div class="card">
          <h3>Yeni IT Talebi</h3>

          <form method="post" enctype="multipart/form-data">
            <div>
              <label class="muted">Konu</label>
              <input name="subject" required>
            </div>

            <div style="margin-top:10px">
              <label class="muted">Öncelik</label>
              <select name="priority">
                {''.join(f"<option>{p}</option>" for p in IT_PRIORITIES)}
              </select>
            </div>

            <div style="margin-top:10px">
              <label class="muted">Açıklama</label>
              <textarea name="description"></textarea>
            </div>

            <div style="margin-top:10px">
              <label class="muted">Ekran Görüntüsü (opsiyonel)</label>
              <input type="file" name="image" accept="image/*">
            </div>

            <div style="margin-top:12px;display:flex;gap:10px;flex-wrap:wrap">
              <button class="btn" type="submit">Talep Oluştur</button>
              <a class="btn2" href="/it/my">İptal</a>
            </div>
          </form>
        </div>
        """
        return render_page("IT Talep Aç", body, user=u)

    # POST -> create
    subject = (request.form.get("subject") or "").strip()
    description = (request.form.get("description") or "").strip()
    priority = (request.form.get("priority") or "Orta").strip()
    if priority not in IT_PRIORITIES:
        priority = "Orta"

    if not subject:
        body = "<div class='card'><div class='pill bad'>Konu boş olamaz.</div><div style='margin-top:10px'><a class='btn2' href='/it/new'>Geri</a></div></div>"
        return render_page("IT Talep Aç", body, user=u)

    image = request.files.get("image")
    image_path = None

    if image and image.filename:
        if not allowed_image(image.filename):
            body = "<div class='card'><div class='pill bad'>Geçersiz resim formatı. (png/jpg/jpeg/webp)</div><div style='margin-top:10px'><a class='btn2' href='/it/new'>Geri</a></div></div>"
            return render_page("IT Talep Aç", body, user=u)

        os.makedirs(UPLOAD_DIR, exist_ok=True)
        filename = secure_filename(image.filename)
        filename = f"{u['id']}_{int(time.time())}_{filename}"
        save_path = os.path.join(UPLOAD_DIR, filename)
        image.save(save_path)
        image_path = f"uploads/it/{filename}"

    now = datetime.now().isoformat(timespec="seconds")
    status = "Açık"

    tid = exec_sql("""
        INSERT INTO it_tickets (user_id, subject, description, priority, status, image_path, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (u["id"], subject, description, priority, status, image_path, now, now))

    # --- MAIL: Yeni IT Talebi -> IT ekibi ---
    try:
        it_emails = get_it_emails()
        if it_emails:
            status = "Açık"
            mail_html = ik_mail_template(
                title="Yeni IT Talebi",
                intro="Yeni bir IT talebi oluşturuldu.",
                rows=[
                    ("Talep No", str(tid)),
                    ("Talep Eden", html_escape(u["full_name"])),
                    ("Konu", html_escape(subject)),
                    ("Öncelik", html_escape(priority)),
                    ("Durum", html_escape(status)),
                ],
                button_text="Talebi Görüntüle",
                button_url=f"{public_base_url()}/it/{tid}",
                footer="İK Portal - IT"
            )
            send_mail_many(it_emails, "[IK][IT] Yeni IT Talebi", mail_html)
    except Exception:
        pass

    return redirect("/it/my")

@app.route("/it/<int:tid>", methods=["GET", "POST"])
@login_required
def it_detail(tid):
    u = current_user()

    t = query_one("""
        SELECT it.*, us.full_name AS owner_name
        FROM it_tickets it
        JOIN users us ON us.id = it.user_id
        WHERE it.id = ?
    """, (tid,))
    if not t:
        abort(404)

    if not _can_view_ticket(u, t):
        abort(403)

    # POST: Mesaj ekleme
    if request.method == "POST":
        msg = (request.form.get("message") or "").strip()
        if msg:
            now = datetime.now().isoformat(timespec="seconds")

            # Mesajı kaydet
            exec_sql("""
                INSERT INTO it_ticket_messages (ticket_id, author_id, message, created_at)
                VALUES (?, ?, ?, ?)
            """, (tid, u["id"], msg, now))

            # Talebi güncelle
            exec_sql(
                "UPDATE it_tickets SET updated_at=? WHERE id=?",
                (now, tid)
            )

            # --- MAIL bildirimi (TEK ve TEMİZ BLOK) ---
            try:
                if u["role"] == ROLE_IT:
                    # IT yazdı -> talep sahibine
                    owner_email = get_user_email(int(t["user_id"]))
                    if owner_email:
                        mail_html = ik_mail_template(
                            title="IT Talebinize Yanıt Var",
                            intro="IT tarafından talebinize yeni bir mesaj yazıldı.",
                            rows=[
                                ("Talep No", str(tid)),
                                ("Konu", html_escape(t["subject"])),
                                ("Mesaj", nl2br(msg)),
                            ],
                            button_text="Talebi Aç",
                            button_url=f"{public_base_url()}/it/{tid}",
                            footer="İK Portal - IT"
                        )
                        send_mail(
                            owner_email,
                            "[IK][IT] Talebinize Yanıt",
                            mail_html
                        )
                else:
                    # Personel yazdı -> IT ekibine
                    it_emails = get_it_emails()
                    if it_emails:
                        mail_html = ik_mail_template(
                            title="IT Talebine Yeni Mesaj",
                            intro="Bir IT talebine yeni mesaj eklendi.",
                            rows=[
                                ("Talep No", str(tid)),
                                ("Talep Eden", html_escape(t["owner_name"])),
                                ("Mesaj", nl2br(msg)),
                            ],
                            button_text="Talebi Aç",
                            button_url=f"{public_base_url()}/it/{tid}",
                            footer="İK Portal - IT"
                        )
                        send_mail_many(
                            it_emails,
                            "[IK][IT] Yeni Mesaj",
                            mail_html
                        )
            except Exception:
                pass

        return redirect(f"/it/{tid}")

    # Mesajları çek
    msgs = query_all("""
        SELECT m.*, u.full_name AS author_name, u.role AS author_role
        FROM it_ticket_messages m
        JOIN users u ON u.id = m.author_id
        WHERE m.ticket_id = ?
        ORDER BY m.id ASC
    """, (tid,))

    msgs_html = ""
    for m in msgs:
        rolepill = f"<span class='pill'>{html_escape(m['author_role'])}</span>"
        msgs_html += f"""
        <div class="note" style="margin-top:10px">
          <div style="display:flex; gap:10px; align-items:center; flex-wrap:wrap;">
            <b>{html_escape(m['author_name'])}</b> {rolepill}
            <span class="muted">{html_escape((m['created_at'] or '')[:19].replace("T"," "))}</span>
          </div>
          <div style="margin-top:8px">{nl2br(m['message'])}</div>
        </div>
        """

    if not msgs_html:
        msgs_html = "<div class='muted'>Henüz mesaj yok.</div>"

    # Görsel
    img_html = ""
    image_path = t["image_path"] if t["image_path"] else ""
    if image_path:
        safe_path = html_escape(image_path)
        img_html = f"""
        <div style="margin-top:12px">
          <div class="muted" style="margin-bottom:6px">Ekran Görüntüsü</div>
          <a href="/uploads/{safe_path.replace('uploads/','')}" target="_blank">
            <img src="/uploads/{safe_path.replace('uploads/','')}"
                 alt="Ekran Görüntüsü"
                 style="max-width:100%;border-radius:12px;border:1px solid rgba(0,0,0,.08)">
          </a>
        </div>
        """

    # IT ise durum güncelleme kutusu
    it_admin_box = ""
    if u["role"] == ROLE_IT:
        status_opts = "".join([
            f"<option value='{html_escape(s)}' {'selected' if (t['status']==s) else ''}>{html_escape(s)}</option>"
            for s in IT_STATUSES
        ])
        it_admin_box = f"""
        <div class="note" style="margin-top:14px">
          <b>IT İşlemi</b>
          <form method="post" action="/it/{tid}/update" style="margin-top:10px">
            <div class="row">
              <div>
                <label class="muted">Durum</label>
                <select name="status">{status_opts}</select>
              </div>
              <div>
                <label class="muted">Durum Notu (IT)</label>
                <input name="status_note" value="{html_escape(t['status_note'] or '')}" placeholder="Kısa not...">
              </div>
            </div>
            <div style="margin-top:10px;display:flex;gap:10px;flex-wrap:wrap">
              <button class="btn" type="submit">Güncelle</button>
              <a class="btn2" href="/it/admin">Listeye Dön</a>
            </div>
          </form>
        </div>
        """

    body = f"""
    <div class="card">
      <h1 class="h1">IT Talep #{t['id']}</h1>
      <div class="two" style="margin-top:10px">
        <div class="note">
          <div class="muted">Talep Sahibi</div>
          <div><b>{html_escape(t['owner_name'])}</b></div>
          <div class="muted" style="margin-top:6px">Oluşturma: {html_escape((t['created_at'] or '')[:19].replace("T"," "))}</div>
          <div class="muted">Güncelleme: {html_escape((t['updated_at'] or '')[:19].replace("T"," ")) if t['updated_at'] else "—"}</div>
        </div>
        <div class="note">
          <div class="muted">Durum / Öncelik</div>
          <div style="margin-top:6px">Durum: {format_status_pill(t['status'])}</div>
          <div style="margin-top:6px">Öncelik: <span class="pill">{html_escape(t['priority'])}</span></div>
          <div style="margin-top:6px" class="muted">IT Notu: {html_escape(t['status_note'] or '-')}</div>
        </div>
      </div>

      <div class="note" style="margin-top:14px">
        <div class="muted">Konu</div>
        <div style="font-weight:800; margin-top:6px">{html_escape(t['subject'])}</div>
        <div class="muted" style="margin-top:10px">Açıklama</div>
        <div style="margin-top:6px">{nl2br(t['description'] or '-')}</div>
        {img_html}
      </div>

      {it_admin_box}

      <div class="note" style="margin-top:14px">
        <b>Mesajlar</b>
        <div style="margin-top:8px">{msgs_html}</div>

        <form method="post" style="margin-top:14px">
          <label class="muted">Yeni Mesaj</label>
          <textarea name="message" placeholder="Mesaj yaz..."></textarea>
          <div style="margin-top:10px;display:flex;gap:10px;flex-wrap:wrap">
            <button class="btn" type="submit">Mesaj Gönder</button>
            <a class="btn2" href="/it/my">Geri</a>
          </div>
        </form>
      </div>
    </div>
    """
    return render_page(f"IT Talep {t['id']}", body, user=u)


@app.route("/it/<int:tid>/update", methods=["POST"])
@role_required(ROLE_IT)
def it_update(tid):
    u = current_user()

    t = query_one("""
        SELECT it.*, us.full_name AS owner_name
        FROM it_tickets it
        JOIN users us ON us.id = it.user_id
        WHERE it.id=?
    """, (tid,))
    if not t:
        abort(404)

    new_status = (request.form.get("status") or "").strip()
    status_note = (request.form.get("status_note") or "").strip()

    if new_status not in IT_STATUSES:
        new_status = t["status"]

    now = datetime.now().isoformat(timespec="seconds")
    closed_at = now if new_status == "Kapalı" else None

    exec_sql("""
        UPDATE it_tickets
        SET status=?, status_note=?, updated_at=?, closed_at=COALESCE(?, closed_at)
        WHERE id=?
    """, (new_status, status_note, now, closed_at, tid))

    # --- MAIL: Durum güncelleme -> talep sahibine ---
    try:
        owner_email = get_user_email(int(t["user_id"]))
        if owner_email:
            mail_html = ik_mail_template(
                title="IT Talebiniz Güncellendi",
                intro="IT talebinizin durumu güncellendi.",
                rows=[
                    ("Talep No", str(tid)),
                    ("Konu", html_escape(t["subject"])),
                    ("Eski Durum", html_escape(t["status"])),
                    ("Yeni Durum", html_escape(new_status)),
                    ("Talep Açıklaması", nl2br(t["description"] or "-")),
                    ("IT Açıklaması", nl2br(status_note or "-")),
                ],
                button_text="Talebi Görüntüle",
                button_url=f"{public_base_url()}/it/{tid}",
                footer="İK Portal - IT"
            )
            send_mail(owner_email, "[IK][IT] Talep Durumu Güncellendi", mail_html)
    except Exception:
        pass

    return redirect(f"/it/{tid}")


@app.route("/it/my")
@login_required
def it_my():
    u = current_user()

    rows = query_all("""
        SELECT id, subject, priority, status, created_at
        FROM it_tickets
        WHERE user_id = ?
        ORDER BY id DESC
    """, (u["id"],))

    trs = ""
    for r in rows:
        trs += f"""
        <tr>
          <td>{r['id']}</td>
          <td>{html_escape(r['subject'])}</td>
          <td><span class="pill">{html_escape(r['priority'])}</span></td>
          <td>{format_status_pill(r['status'])}</td>
          <td class="muted">{html_escape((r['created_at'] or '')[:19].replace("T"," "))}</td>
          <td><a class="btn2" href="/it/{r['id']}">Detay</a></td>
        </tr>
        """

    if not trs:
        trs = "<tr><td colspan='6' class='muted'>Henüz IT talebiniz yok.</td></tr>"

    body = f"""
    <div class="card">
      <h1 class="h1">IT Taleplerim</h1>
      <div style="margin-bottom:12px">
        <a class="btn" href="/it/new">Yeni IT Talebi Aç</a>
      </div>
      <table>
        <thead>
          <tr><th>ID</th><th>Konu</th><th>Öncelik</th><th>Durum</th><th>Oluşturma</th><th></th></tr>
        </thead>
        <tbody>{trs}</tbody>
      </table>
      <div class="note" style="margin-top:14px">
        <div class="muted">Bu ekranda sadece size ait IT talepleri listelenir.</div>
      </div>
    </div>
    """
    return render_page("IT Taleplerim", body, user=u)


@app.route("/it/admin")
@role_required(ROLE_IT)
def it_admin():
    rows = query_all("""
        SELECT it.id, it.subject, it.priority, it.status,
               it.created_at, it.updated_at,
               u.full_name AS owner_name
        FROM it_tickets it
        JOIN users u ON u.id = it.user_id
        ORDER BY it.id DESC
    """)

    trs = ""
    for r in rows:
        trs += f"""
        <tr>
          <td>{r['id']}</td>
          <td>{html_escape(r['subject'])}</td>
          <td>{html_escape(r['owner_name'])}</td>
          <td><span class="pill">{html_escape(r['priority'])}</span></td>
          <td>{format_status_pill(r['status'])}</td>
          <td class="muted">{html_escape((r['created_at'] or '')[:19].replace("T"," "))}</td>
          <td><a class="btn2" href="/it/{r['id']}">Detay</a></td>
        </tr>
        """

    if not trs:
        trs = "<tr><td colspan='7' class='muted'>Henüz talep yok.</td></tr>"

    body = f"""
    <div class="card">
      <h1 class="h1">IT Talep Yönetimi</h1>
      <div style="margin-bottom:12px;display:flex;gap:10px;flex-wrap:wrap">
        <a class="btn btn-excel" href="/it/admin.xlsx">Excel İndir</a>
      </div>

      <table>
        <thead>
          <tr><th>ID</th><th>Konu</th><th>Açan</th><th>Öncelik</th><th>Durum</th><th>Oluşturma</th><th></th></tr>
        </thead>
        <tbody>{trs}</tbody>
      </table>

      <div class="note" style="margin-top:14px">
        <div class="muted">Bu ekran sadece IT ekibi tarafından görüntülenebilir.</div>
      </div>
    </div>
    """
    return render_page("IT Talep Yönetimi", body, user=current_user())

@app.route("/it/admin.xlsx")
@role_required(ROLE_IT)
def it_admin_xlsx():
    rows = query_all("""
        SELECT it.id, u.full_name AS owner_name, u.username AS owner_username,
               it.subject, it.priority, it.status, it.status_note,
               it.created_at, it.updated_at, it.closed_at
        FROM it_tickets it
        JOIN users u ON u.id = it.user_id
        ORDER BY it.id DESC
    """)

    wb = Workbook()
    ws = wb.active
    ws.title = "IT Talepleri"

    ws.append([
        "ID", "Açan", "Kullanıcı", "Konu", "Öncelik", "Durum", "IT Notu",
        "Oluşturma", "Güncelleme", "Kapanış"
    ])

    for r in rows:
        ws.append([
            r["id"],
            r["owner_name"],
            r["owner_username"],
            r["subject"],
            r["priority"],
            r["status"],
            r["status_note"] or "",
            (r["created_at"] or "")[:19].replace("T", " "),
            (r["updated_at"] or "")[:19].replace("T", " ") if r["updated_at"] else "",
            (r["closed_at"] or "")[:19].replace("T", " ") if r["closed_at"] else "",
        ])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"it_talepleri_{date.today().isoformat()}.xlsx"
    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )

@app.route("/admin/announcements")
def admin_announcements():
    if not current_user():
        return redirect(url_for("login", next=request.path))

    if not is_admin_user():
        abort(403)

    conn = db_connect()
    rows = conn.execute("""
        SELECT a.*,
               u1.username AS created_by_name,
               u2.username AS updated_by_name
        FROM announcements a
        LEFT JOIN users u1 ON u1.id = a.created_by
        LEFT JOIN users u2 ON u2.id = a.updated_by
        ORDER BY a.id DESC
    """).fetchall()

    html = """
    <div class="card">
      <div style="display:flex;align-items:center;justify-content:space-between;gap:12px;">
        <div>
          <div style="font-size:18px;font-weight:800;">Duyurular</div>
          <div class="muted">Popup + ana sayfa duyuruları buradan yönetilir.</div>
        </div>
        <a class="btn2" href="/admin/announcements/new">+ Yeni Duyuru</a>
      </div>
    </div>

    <div class="card" style="margin-top:12px;">
      <table class="table">
        <thead>
          <tr>
            <th>ID</th>
            <th>Başlık</th>
            <th>Durum</th>
            <th>Oluşturan</th>
            <th>Tarih</th>
            <th style="width:260px;">İşlem</th>
          </tr>
        </thead>
        <tbody>
    """

    for r in rows:
        is_active = int(r["is_active"]) == 1
        status_html = "<span class='pill ok'>Aktif</span>" if is_active else "<span class='pill bad'>Pasif</span>"
        created_by = (r["created_by_name"] or "-")
        created_at = (r["created_at"] or "-")
        title = (r["title"] or "").replace("<", "&lt;").replace(">", "&gt;")

        toggle_txt = "Pasif Yap" if is_active else "Aktif Yap"

        html += f"""
          <tr>
            <td>{r["id"]}</td>
            <td><b>{title}</b></td>
            <td>{status_html}</td>
            <td>{created_by}</td>
            <td>{created_at}</td>
            <td>
              <a class="btn2" href="/admin/announcements/{r['id']}/edit">Düzenle</a>
              <a class="btn2" href="/admin/announcements/{r['id']}/toggle">{toggle_txt}</a>
              <a class="btn2" href="/admin/announcements/{r['id']}/delete" onclick="return confirm('Silinsin mi?');">Sil</a>
            </td>
          </tr>
        """

    html += """
        </tbody>
      </table>
    </div>
    """

    return render_page("Duyurular", html)

@app.route("/admin/announcements/new", methods=["GET", "POST"])
def admin_announcement_new():
    if not is_admin_user():
        abort(403)

    if request.method == "POST":
        title = (request.form.get("title") or "").strip()
        body = (request.form.get("body") or "").strip()
        is_active = 1 if (request.form.get("is_active") == "1") else 0

        if not title or not body:
            return render_page("Yeni Duyuru", """
            <div class="card"><div class="pill bad">Başlık ve açıklama zorunlu.</div></div>
            <div style="height:10px"></div>
            <a class="btn2" href="/admin/announcements/new">Geri</a>
            """)

        image_path = None
        file = request.files.get("image")
        if file and file.filename:
            fn = secure_filename(file.filename)
            ext = fn.rsplit(".", 1)[-1].lower() if "." in fn else ""
            if ext not in ("png", "jpg", "jpeg", "webp"):
                return render_page("Yeni Duyuru", """
                <div class="card"><div class="pill bad">Resim formatı desteklenmiyor (png/jpg/jpeg/webp).</div></div>
                <div style="height:10px"></div>
                <a class="btn2" href="/admin/announcements/new">Geri</a>
                """)
            new_name = f"ann_{int(time.time())}_{secrets.token_hex(4)}.{ext}"
            save_path = os.path.join(ANN_UPLOAD_DIR, new_name)
            file.save(save_path)
            image_path = f"uploads/announcements/{new_name}"

        conn = db_connect()
        conn.execute("""
            INSERT INTO announcements (title, body, image_path, is_active, created_at, created_by)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (
            title,
            body,
            image_path,
            is_active,
            datetime.now().strftime("%Y-%m-%d %H:%M"),
            session.get("uid")
        ))
        conn.commit()

        return redirect("/admin/announcements")

    # GET form
    html = """
    <div class="card" style="max-width:760px;margin:0 auto;">
      <div style="font-size:18px;font-weight:800;">Yeni Duyuru</div>
      <div class="muted">Kaydettikten sonra popup/anasayfada görünecek.</div>

      <form method="POST" enctype="multipart/form-data" style="margin-top:12px;">
        <label class="muted">Başlık</label>
        <input class="inp" name="title" placeholder="Örn: Sistem Bakımı" />

        <div style="height:10px"></div>

        <label class="muted">Açıklama</label>
        <textarea class="inp" name="body" rows="6" placeholder="Detay..."></textarea>

        <div style="height:10px"></div>

        <label class="muted">Resim (opsiyonel)</label>
        <input class="inp" type="file" name="image" accept=".png,.jpg,.jpeg,.webp" />

        <div style="height:10px"></div>

        <label style="display:flex;gap:8px;align-items:center;">
          <input type="checkbox" name="is_active" value="1" checked />
          <span>Aktif</span>
        </label>

        <div style="height:14px"></div>

        <button class="btn2" type="submit">Kaydet</button>
        <a class="btn2" href="/admin/announcements" style="margin-left:8px;">Vazgeç</a>
      </form>
    </div>
    """
    return render_page("Yeni Duyuru", html)

@app.route("/announcement/<int:aid>")
@login_required
def announcement_detail(aid):

    row = query_one("""
        SELECT id, title, body, image_path, created_at
        FROM announcements
        WHERE id=? AND is_active=1
    """, (aid,))

    if not row:
        return redirect("/")

    img_html = ""
    if row["image_path"]:
        safe_img = html_escape(row["image_path"]).replace("uploads/", "")
        img_html = f"""
        <div style="margin-top:14px">
          <img src="/uploads/{safe_img}"
               style="max-width:100%;border-radius:16px;border:1px solid rgba(148,163,184,.20)">
        </div>
        """

    created = html_escape((row["created_at"] or "")[:19].replace("T", " "))

    body = f"""
    <div class="card" style="max-width:900px;margin:0 auto">
        <h2 class="h1">{html_escape(row["title"] or "")}</h2>
        <div class="muted" style="margin-top:6px">{created}</div>
        <div style="margin-top:14px">
            {nl2br(row["body"] or "")}
        </div>
        {img_html}
        <div style="margin-top:20px">
            <a class="btn2" href="/">← Ana Sayfa</a>
        </div>
    </div>
    """

    return render_page("Duyuru", body)

@app.route("/reminders", methods=["GET", "POST"])
@login_required
def reminders_page():
    app.logger.warning("DEBUG ROLE=%s UID=%s", session.get("role"), session.get("uid"))
    if request.method == "POST":
        title = (request.form.get("title") or "").strip()
        due_date = (request.form.get("due_date") or "").strip()  # YYYY-MM-DD
        remind_days = (request.form.get("remind_days") or "30,7,1,0").strip()
        target_roles = (request.form.get("target_roles") or "").strip()
        target_emails = (request.form.get("target_emails") or "").strip()
        scope = (request.form.get("scope") or "private").strip()
        created_by = session.get("uid")

        role = session.get("role")

        # 🔒 Personnel güvenlik zorlaması
        if role == ROLE_PERSONNEL:
            scope = "private"
            target_roles = ""

        created_by = session.get("uid")

        if not title or not due_date:
            return render_page("Hatırlatmalar", "<p>Başlık ve tarih zorunlu.</p><p><a href='/reminders'>Geri</a></p>")

        if (not target_roles) and (not target_emails):
            return render_page("Hatırlatmalar", "<p>Hedef Roller veya Ek E-postalar alanlarından en az biri dolu olmalı.</p><p><a href='/reminders'>Geri</a></p>")

        exec_sql("""
            INSERT INTO reminders 
            (title, due_date, remind_days, target_roles, target_emails, is_active, created_at, created_by, scope)
            VALUES (?, ?, ?, ?, ?, 1, ?, ?, ?)
        """, (
            title,
            due_date,
            remind_days,
            target_roles,
            target_emails,
            datetime.now().isoformat(timespec="seconds"),
            created_by,
            scope
        ))

        return redirect("/reminders")

    current_user_id = session.get("uid")
    current_role = session.get("role")

    if current_role == ROLE_OWNER:
        rows = query_all("""
            SELECT * FROM reminders
            ORDER BY date(due_date) ASC
        """)
    else:
        rows = query_all("""
            SELECT * FROM reminders
            WHERE
                created_by = ?
                OR scope = 'global'
            ORDER BY date(due_date) ASC
        """, (current_user_id,))

    trs = ""
    for r in rows:
        status_badge = "Aktif" if r["is_active"] else "Pasif"
        toggle_txt = "Pasif Yap" if r["is_active"] else "Aktif Yap"
        trs += f"""
        <tr>
          <td>{r['id']}</td>
          <td>{html_escape(r['title'])}</td>
          <td>{html_escape(r['due_date'])}</td>
          <td>{html_escape(r['remind_days'])}</td>
          <td>{html_escape(r['target_roles'])}</td>
          <td>{html_escape(r['target_emails'] or '')}</td>
          <td>{status_badge}</td>
          <td style="white-space:nowrap">
            <a class="btn" href="/reminders/{r['id']}/toggle">{toggle_txt}</a>
            <a class="btn danger" href="/reminders/{r['id']}/delete" onclick="return confirm('Silinsin mi?')">Sil</a>
          </td>
        </tr>
        """

    role = session.get("role")

    target_roles_html = ""
    if role != ROLE_PERSONNEL:
        target_roles_html = """
        <div>
          <label>Hedef Roller (örn: owner,accounting)</label>
          <input name="target_roles" value="" placeholder="örn: owner,accounting" />
        </div>
        """

    if role == ROLE_PERSONNEL:
        scope_html = """
        <div>
          <label>Kapsam</label>
          <input type="hidden" name="scope" value="private" />
          <div class="muted" style="padding:10px 12px;border:1px solid #e5e7eb;border-radius:10px;background:#f9fafb">
            Sadece Ben (Personel için sabit)
          </div>
        </div>
        """
    else:
        scope_html = """
        <div>
          <label>Kapsam</label>
          <select name="scope">
            <option value="private">Sadece Ben</option>
            <option value="team">Ekibim</option>
            <option value="global">Genel (Herkes)</option>
          </select>
        </div>
        """

    body = f"""
    <h2>Hatırlatmalar (Süre Takibi)</h2>

    <div class="card" style="margin-bottom:16px">
      <h3>Yeni Hatırlatma Ekle</h3>
      <form method="post" style="display:grid;grid-template-columns:1fr 220px;gap:12px;align-items:end">
        <div>
          <label>Başlık (Örn: 34ABC123 - Seyrüsefer)</label>
          <input name="title" required placeholder="Araç/konu - işlem" />
        </div>

        {scope_html}

        <div>
          <label>Bitiş / Son Tarih</label>
          <input type="date" name="due_date" required />
        </div>

        <div>
          <label>Kaç gün kala mail? (örn: 30,7,1,0)</label>
          <input name="remind_days" value="30,7,1,0" />
        </div>

        {target_roles_html}

        <div style="grid-column:1 / span 2">
          <label>Ek E-postalar (opsiyonel, virgülle) (örn: a@x.com,b@y.com)</label>
          <input name="target_emails" placeholder="opsiyonel" />
        </div>

        <div style="grid-column:1 / span 2">
          <button class="btn primary" type="submit">Kaydet</button>
        </div>
      </form>
    </div>

    <div class="card">
      <h3>Kayıtlar</h3>
      <table class="table" style="width:100%">
        <thead>
          <tr>
            <th>ID</th><th>Başlık</th><th>Tarih</th><th>Günler</th><th>Roller</th><th>Ek Mail</th><th>Durum</th><th>İşlem</th>
          </tr>
        </thead>
        <tbody>
          {trs if trs else "<tr><td colspan='8' class='muted'>Henüz kayıt yok.</td></tr>"}
        </tbody>
      </table>
    </div>
    """
    return render_page("Hatırlatmalar", body)

@app.route("/admin/announcements/<int:aid>/edit", methods=["GET", "POST"])
def admin_announcement_edit(aid):
    if not session.get("uid"):
        return redirect(url_for("login", next=request.path))
    if not is_admin_user():
        abort(403)

    conn = db_connect()
    row = conn.execute("SELECT * FROM announcements WHERE id = ?", (aid,)).fetchone()
    if not row:
        abort(404)

    if request.method == "POST":
        title = (request.form.get("title") or "").strip()
        body = (request.form.get("body") or "").strip()
        is_active = 1 if (request.form.get("is_active") == "1") else 0

        if not title or not body:
            return render_page("Duyuru Düzenle", "<div class='card'><div class='pill bad'>Başlık ve açıklama zorunlu.</div></div>")

        image_path = row["image_path"]
        file = request.files.get("image")
        if file and file.filename:
            fn = secure_filename(file.filename)
            ext = fn.rsplit(".", 1)[-1].lower() if "." in fn else ""
            if ext not in ("png", "jpg", "jpeg", "webp"):
                return render_page("Duyuru Düzenle", "<div class='card'><div class='pill bad'>Resim formatı desteklenmiyor.</div></div>")
            new_name = f"ann_{int(time.time())}_{secrets.token_hex(4)}.{ext}"
            save_path = os.path.join(ANN_UPLOAD_DIR, new_name)
            file.save(save_path)
            image_path = f"/uploads/announcements/{new_name}"

        conn.execute("""
            UPDATE announcements
            SET title = ?, body = ?, image_path = ?, is_active = ?,
                updated_at = ?, updated_by = ?
            WHERE id = ?
        """, (
            title,
            body,
            image_path,
            is_active,
            datetime.now().strftime("%Y-%m-%d %H:%M"),
            session.get("uid"),
            aid
        ))
        conn.commit()

        return redirect("/admin/announcements")

    # GET form
    checked = "checked" if int(row["is_active"]) == 1 else ""
    title = (row["title"] or "").replace('"', "&quot;")
    body = (row["body"] or "")

    img_html = ""
    if row["image_path"]:
        img_html = f"""
        <div style="height:10px"></div>
        <div class="muted">Mevcut Resim</div>
        <img src="{row['image_path']}" style="max-width:100%;border-radius:12px;border:1px solid var(--line);" />
        """

    html = f"""
    <div class="card" style="max-width:760px;margin:0 auto;">
      <div style="font-size:18px;font-weight:800;">Duyuru Düzenle (#{row['id']})</div>

      <form method="POST" enctype="multipart/form-data" style="margin-top:12px;">
        <label class="muted">Başlık</label>
        <input class="inp" name="title" value="{title}" />

        <div style="height:10px"></div>

        <label class="muted">Açıklama</label>
        <textarea class="inp" name="body" rows="6">{body}</textarea>

        {img_html}

        <div style="height:10px"></div>

        <label class="muted">Resim Değiştir (opsiyonel)</label>
        <input class="inp" type="file" name="image" accept=".png,.jpg,.jpeg,.webp" />

        <div style="height:10px"></div>

        <label style="display:flex;gap:8px;align-items:center;">
          <input type="checkbox" name="is_active" value="1" {checked} />
          <span>Aktif</span>
        </label>

        <div style="height:14px"></div>

        <button class="btn2" type="submit">Kaydet</button>
        <a class="btn2" href="/admin/announcements" style="margin-left:8px;">Geri</a>
      </form>
    </div>
    """
    return render_page("Duyuru Düzenle", html)


@app.route("/admin/announcements/<int:aid>/toggle")
def admin_announcement_toggle(aid):
    if not session.get("uid"):
        return redirect(url_for("login", next=request.path))
    if not is_admin_user():
        abort(403)

    conn = db_connect()
    row = conn.execute("SELECT is_active FROM announcements WHERE id = ?", (aid,)).fetchone()
    if not row:
        abort(404)

    new_val = 0 if int(row["is_active"]) == 1 else 1
    conn.execute("""
        UPDATE announcements
        SET is_active = ?, updated_at = ?, updated_by = ?
        WHERE id = ?
    """, (
        new_val,
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        session.get("uid"),
        aid
    ))
    conn.commit()

    return redirect("/admin/announcements")


@app.route("/admin/announcements/<int:aid>/delete")
def admin_announcement_delete(aid):
    if not session.get("uid"):
        return redirect(url_for("login", next=request.path))
    if not is_admin_user():
        abort(403)

    conn = db_connect()
    conn.execute("DELETE FROM announcements WHERE id = ?", (aid,))
    conn.commit()

    return redirect("/admin/announcements")

@app.route("/reminders/<int:rid>/toggle")
@login_required
def reminders_toggle(rid: int):
    r = query_one("SELECT id, is_active FROM reminders WHERE id=?", (rid,))
    if not r:
        abort(404)
    new_val = 0 if int(r["is_active"]) == 1 else 1
    exec_sql("UPDATE reminders SET is_active=? WHERE id=?", (new_val, rid))
    return redirect("/reminders")


@app.route("/reminders/<int:rid>/delete")
@login_required
def reminders_delete(rid: int):
    exec_sql("DELETE FROM reminder_notifications WHERE reminder_id=?", (rid,))
    exec_sql("DELETE FROM reminders WHERE id=?", (rid,))
    return redirect("/reminders")


@app.route("/logout")
def logout():
    session.clear()
    return redirect("/login")

@app.route("/uploads/<path:filename>")
@login_required
def uploads_serve(filename):
    """
    Güvenli şekilde uploads altından dosya servis eder.
    Sadece uploads/ içinden izin verir (path traversal engeli).
    """
    base_dir = os.path.join(os.path.dirname(__file__), "uploads")
    safe_path = os.path.normpath(filename).lstrip("/")

    full_path = os.path.join(base_dir, safe_path)

    # Path traversal engeli: full_path mutlaka base_dir altında olmalı
    if not os.path.abspath(full_path).startswith(os.path.abspath(base_dir) + os.sep):
        abort(403)

    if not os.path.exists(full_path) or not os.path.isfile(full_path):
        abort(404)

    return send_file(full_path)


@app.route("/")
@login_required
def index():
    u = current_user()
    today = date.today().isoformat()
    year = date.today().year
    used = used_annual_leave_days(int(u["id"]), year)
    remaining = remaining_annual_leave_days(u, year)

    my_leave = query_all("""
      SELECT * FROM leave_requests
      WHERE user_id=?
      ORDER BY id DESC
      LIMIT 5
    """, (u["id"],))

    my_att = query_all("""
      SELECT * FROM attendance
      WHERE user_id=? AND tarih=?
      ORDER BY id ASC
    """, (u["id"], today))

    att_html = "<div class='muted'>Bugün kayıt yok.</div>"
    if my_att:
        rows = "".join([f"<tr><td>{html_escape(r['saat'])}</td><td>{html_escape(TUR_LABEL.get(r['tur'], r['tur']))}</td></tr>" for r in my_att])
        att_html = f"<table><thead><tr><th>Saat</th><th>İşlem</th></tr></thead><tbody>{rows}</tbody></table>"

    leave_html = "<div class='muted'>Henüz izin talebiniz yok.</div>"
    if my_leave:
        rows = "".join([
            f"<tr>"
            f"<td>{html_escape(r['start_date'])}</td>"
            f"<td>{html_escape(r['end_date'])}</td>"
            f"<td>{html_escape(r['leave_type'])}</td>"
            f"<td>{('Yarım Gün (' + html_escape(r['half_day_part'] or '-') + ')') if int(r['is_half_day'] or 0)==1 else 'Tam Gün'}</td>"
            f"<td>{format_status_pill(r['status'])}</td>"
            f"</tr>"
            for r in my_leave
        ])
        leave_html = (
            "<table><thead><tr>"
            "<th>Başlangıç</th><th>Bitiş</th><th>Tür</th><th>Gün</th><th>Durum</th>"
            "</tr></thead><tbody>"
            f"{rows}"
            "</tbody></table>"
        )

    qr_card = ""
    if u["can_qr"] == 1:
        qr_card = """
        <div class="card">
          <h2 class="h1">Mesai (QR)</h2>
          <p class="muted">Telefon kameranızla okutmak için QR ekranını açın.</p>
          <a class="btn" href="/qr">QR ile Mesai</a>
        </div>
        """

    leave_buttons = ""
    if u["role"] != ROLE_OWNER:
        leave_buttons = """
          <a class="btn" href="/leave/new">İzin Talep Et</a>
          <a class="btn2" href="/leave/my">İzinlerim</a>
        """

    # ✅ DUYURULAR (Ana Sayfa - herkes görür)
    # Not: tablo/kolon adlarını kendi modülündeki isimlerle aynı yap:
    # Örn tablo: announcements / duyurular  | kolon: title/body/image_path/is_active/created_at
    ann_rows = []
    try:
        ann_rows = query_all("""
            SELECT id, title, body, image_path, created_at
            FROM announcements
            WHERE is_active=1
            ORDER BY id DESC
            LIMIT 5
        """)
    except Exception:
        ann_rows = []

    ann_html = ""
    if ann_rows:
        cards = ""
        for a in ann_rows:
            img = ""
            if a["image_path"]:
                # image_path örn: "uploads/duyuru/abc.jpg" veya "uploads/it/..."
                safe_img = html_escape(a["image_path"]).replace("uploads/", "")
                img = f"""
                <div style="margin-top:10px">
                  <a href="/uploads/{safe_img}" target="_blank">
                    <img src="/uploads/{safe_img}"
                         style="max-width:100%;border-radius:14px;border:1px solid rgba(148,163,184,.20)">
                  </a>
                </div>
                """

            created = html_escape((a["created_at"] or "")[:19].replace("T", " "))
            is_new = False
            try:
                cdt = (a["created_at"] or "")[:10]
                if cdt:
                    cdate = datetime.strptime(cdt, "%Y-%m-%d").date()
                    is_new = (date.today() - cdate).days <= 2
            except Exception:
                is_new = False
            cards += f"""
            <div class="ann-item" style="margin-top:10px">
              <div class="ann-top">
                <div class="ann-title" style="font-weight:900">
                  <a href="/announcement/{a['id']}">
                    {html_escape(a["title"] or "")}
                  </a>
                </div>
                  <div class="ann-badges">
                    {("<span class='ann-badge new' style='display:inline-block;background:linear-gradient(135deg,#ff3b3b,#ff0000);color:#fff;border:0;font-weight:900;box-shadow:0 0 12px rgba(255,0,0,.85);padding:4px 10px;border-radius:999px;letter-spacing:.4px;line-height:1;margin-right:8px;vertical-align:middle'>YENİ</span>" if is_new else "")}
                    <span class="ann-badge" style="display:inline-block;line-height:1;vertical-align:middle">{created}</span>
                  </div>
              </div>

              <div style="margin-top:10px;display:flex;gap:20px;align-items:flex-start;flex-wrap:wrap">

                <div style="flex:1;min-width:260px">
                  <div style="margin-top:8px">
                    {nl2br((a["body"] or "")[:280] + ("..." if len(a["body"] or "") > 280 else ""))}
                  </div>

                  <div style="margin-top:12px">
                    <a class="btn2" href="/announcement/{a['id']}">Devamını gör</a>
                  </div>
                </div>

                {"<div style='width:320px;max-width:100%'>" + img + "</div>" if a["image_path"] else ""}

              </div>
            </div>
            """
        ann_html = f"""
        <div class="card" style="margin-top:14px">
          <h2 class="h1">Duyurular</h2>
          {cards}
        </div>
        """
    dino_html = """
    <div class="card" style="margin-top:14px">
      <h2 class="h1">Dino Run</h2>
      <div class="muted">Boşluk (Space) veya tıkla: Zıpla • Yeniden başlat: R</div>

      <div style="margin-top:12px;display:flex;gap:16px;flex-wrap:wrap;align-items:flex-start">
        <div style="flex:1;min-width:320px">
          <canvas id="dinoCanvas" width="760" height="220"
                  style="width:100%;max-width:760px;background:#f8fafc;border:1px solid rgba(148,163,184,.20);border-radius:14px"></canvas>

          <div style="margin-top:10px;display:flex;gap:10px;flex-wrap:wrap;align-items:center">
            <span class="pill" style="display:inline-flex;gap:8px;align-items:center">
              Skor: <b id="dinoScore">0</b>
            </span>
            <span class="pill" style="display:inline-flex;gap:8px;align-items:center">
              En iyi: <b id="dinoBest">0</b>
            </span>
            <span id="dinoStatus" class="muted"></span>
          </div>
        </div>

        <div style="width:320px;max-width:100%">
          <div class="pill" style="margin-bottom:10px">🏆 Liderlik Tablosu (Top 10)</div>
          <div id="dinoLeaders" class="muted">Yükleniyor...</div>
        </div>
      </div>
    </div>

    <script>
    (function(){
      var canvas = document.getElementById("dinoCanvas");
      if(!canvas) return;
      var ctx = canvas.getContext("2d");

      var scoreEl = document.getElementById("dinoScore");
      var bestEl  = document.getElementById("dinoBest");
      var statusEl = document.getElementById("dinoStatus");
      var leadersEl = document.getElementById("dinoLeaders");

      var W = canvas.width, H = canvas.height;
      var groundY = H - 32;

      var dino = { x: 40, y: groundY - 34, w: 28, h: 34, vy: 0, onGround: true };
      var obstacles = [];
      var t = 0;
      var score = 0;
      var best = 0;
      var speed = 3.2;
      var running = true;
      var lastSpawn = 0;

      function clamp(n, a, b){ return Math.max(a, Math.min(b, n)); }

      function reset(){
        obstacles = [];
        t = 0;
        score = 0;
        speed = 3.2;
        dino.y = groundY - dino.h;
        dino.vy = 0;
        dino.onGround = true;
        running = true;
        statusEl.textContent = "";
      }

      function jump(){
        if(!running) return;
        if(dino.onGround){
          dino.vy = -9.8;
          dino.onGround = false;
        }
      }

      function spawnObstacle(){
        var h = 24 + Math.floor(Math.random()*18);
        var w = 14 + Math.floor(Math.random()*10);
        obstacles.push({ x: W + 10, y: groundY - h, w: w, h: h });
      }

      function collide(a,b){
        return !(a.x+a.w < b.x || a.x > b.x+b.w || a.y+a.h < b.y || a.y > b.y+b.h);
      }

      function draw(){
        // bg
        ctx.clearRect(0,0,W,H);

        // ground
        ctx.beginPath();
        ctx.moveTo(0, groundY+0.5);
        ctx.lineTo(W, groundY+0.5);
        ctx.stroke();

        // dino
        ctx.fillRect(dino.x, dino.y, dino.w, dino.h);

        // obstacles
        for(var i=0;i<obstacles.length;i++){
          var o = obstacles[i];
          ctx.fillRect(o.x, o.y, o.w, o.h);
        }

        // subtle text
        ctx.font = "12px sans-serif";
        ctx.fillText("Space / Click: Jump • R: Restart", 10, 18);
      }

      function update(){
        if(!running){
          draw();
          return;
        }

        t++;

        // speed up slowly
        speed = 3.2 + (t/600);

        // gravity
        dino.vy += 0.55;
        dino.y += dino.vy;

        if(dino.y >= groundY - dino.h){
          dino.y = groundY - dino.h;
          dino.vy = 0;
          dino.onGround = true;
        }

        // spawn
        if(t - lastSpawn > (70 + Math.random()*60)){
          spawnObstacle();
          lastSpawn = t;
        }

        // move obstacles
        for(var i=obstacles.length-1;i>=0;i--){
          obstacles[i].x -= speed;
          if(obstacles[i].x + obstacles[i].w < -10){
            obstacles.splice(i,1);
          }
        }

        // collisions
        for(var j=0;j<obstacles.length;j++){
          if(collide(dino, obstacles[j])){
            gameOver();
            break;
          }
        }

        // score
        score += 1;
        if(score > best) best = score;

        scoreEl.textContent = String(Math.floor(score/5));
        bestEl.textContent  = String(Math.floor(best/5));

        draw();
        requestAnimationFrame(update);
      }

      function gameOver(){
        running = false;
        statusEl.textContent = "💥 Oyun bitti! R ile tekrar başlat.";
        // submit best score
        var bestScore = Math.floor(best/5);

        fetch("/api/dino/submit", {
          method: "POST",
          headers: { "Content-Type": "application/json" },
          body: JSON.stringify({ score: bestScore })
        }).then(function(){ loadLeaders(); }).catch(function(){});
      }

      function loadLeaders(){
        fetch("/api/dino/leaderboard")
          .then(function(r){ return r.json(); })
          .then(function(data){
            var arr = (data && data.leaders) ? data.leaders : [];
            if(!arr.length){
              leadersEl.innerHTML = "<div class='muted'>Henüz skor yok.</div>";
              return;
            }
            var html = "<table style='width:100%'><thead><tr><th>#</th><th>İsim</th><th style='text-align:right'>Skor</th></tr></thead><tbody>";
            for(var i=0;i<arr.length;i++){
              var nm = (arr[i].name || "").replace(/</g,"&lt;").replace(/>/g,"&gt;");
              var sc = arr[i].score;
              html += "<tr><td class='muted'>"+(i+1)+"</td><td>"+nm+"</td><td style='text-align:right'><b>"+sc+"</b></td></tr>";
            }
            html += "</tbody></table>";
            leadersEl.innerHTML = html;
          })
          .catch(function(){
            leadersEl.innerHTML = "<div class='muted'>Liderlik tablosu yüklenemedi.</div>";
          });
      }

      // controls
      document.addEventListener("keydown", function(e){
        if(e.code === "Space"){ e.preventDefault(); jump(); }
        if(e.key === "r" || e.key === "R"){ reset(); requestAnimationFrame(update); }
      });
      canvas.addEventListener("mousedown", function(){ jump(); });
      canvas.addEventListener("touchstart", function(e){ e.preventDefault(); jump(); }, {passive:false});

      // start
      loadLeaders();
      requestAnimationFrame(update);
      setInterval(loadLeaders, 15000);
    })();
    </script>
    """

    body = f"""
    <div class="two">
      <div class="card">
        <h1 class="h1">Hoş geldin, {html_escape(u['full_name'])}</h1>
        <div class="muted">Bugün: {today}</div>
        <div style="margin-top:10px" class="pill">
            <div class="muted">{year} Yıllık İzin: <b>{remaining:.1f}</b> gün kaldı (kullanılan: {used:.1f})</div>
        </div>
        <div style="margin-top:12px;display:flex;gap:10px;flex-wrap:wrap">
          {leave_buttons}
        </div>
      </div>
      {qr_card}
    </div>

    {ann_html}

    <div class="two" style="margin-top:14px">
      <div class="card">
        <h2 class="h1">Bugünkü Mesai Kayıtlarım</h2>
        {att_html}
      </div>
      <div class="card">
        <h2 class="h1">Son İzin Taleplerim</h2>
        {leave_html}
      </div>
    </div>
    
    {dino_html}
    """
    return render_page("Ana Sayfa", body, user=u)


# -----------------------------
# Leave (personnel)
# -----------------------------
@app.route("/leave/my")
@login_required
def leave_my():
    u = current_user()

    if u["role"] == ROLE_OWNER:
        return render_page("İzinlerim", "<div class='card'><div class='pill warn'>Patron için izin talep/izinler ekranı kullanılmaz.</div></div>", user=u)

    rows = query_all("""
      SELECT id, start_date, end_date, leave_type, status, created_at, is_half_day, half_day_part
      FROM leave_requests
      WHERE user_id=?
      ORDER BY id DESC
    """, (u["id"],))

    trs = ""
    for r in rows:
        trs += f"""
        <tr>
          <td>{html_escape(r['start_date'])}</td>
          <td>{html_escape(r['end_date'])}</td>
          <td>
            {html_escape(r['leave_type'])}
            { " (Yarım Gün - " + html_escape(r['half_day_part'] or "") + ")" if int(r['is_half_day'] or 0)==1 else "" }
          </td>
          <td>{format_status_pill(r['status'])}</td>
          <td class="muted">{html_escape((r['created_at'][:19].replace("T"," ")) )}</td>
        </tr>
        """
    if not trs:
        trs = "<tr><td colspan='5' class='muted'>Kayıt yok.</td></tr>"

    body = f"""
    <div class="card">
      <h1 class="h1">İzinlerim</h1>
      <div style="margin-bottom:12px">
        <a class="btn" href="/leave/new">Yeni Talep</a>
      </div>
      <table>
        <thead><tr><th>Başlangıç</th><th>Bitiş</th><th>Süre</th><th>Tür</th><th>Durum</th><th>Oluşturma</th></tr></thead>
        <tbody>{trs}</tbody>
      </table>
    </div>
    """
    return render_page("İzinlerim", body, user=u)

@app.route("/leave/new", methods=["GET", "POST"])
@login_required
def leave_new():
    u = current_user()

    if u["role"] == ROLE_OWNER:
        return render_page("İzin Talep Et", "<div class='card'><div class='pill warn'>Patron için izin talep ekranı kapalıdır.</div></div>", user=u)

    if request.method == "POST":
        app.logger.warning("DEBUG leave_new FORM=%s", dict(request.form))
        start_date = (request.form.get("start_date") or "").strip()
        end_date = (request.form.get("end_date") or "").strip()
        leave_type = (request.form.get("leave_type") or "Yillik").strip()
        reason = (request.form.get("reason") or "").strip()

        if not start_date or not end_date:
            return render_page("İzin Talep Et", "<div class='card'><div class='pill bad'>Tarihleri doldurun.</div></div>", user=u)

        try:
            d1 = datetime.strptime(start_date, "%Y-%m-%d").date()
            d2 = datetime.strptime(end_date, "%Y-%m-%d").date()
            if d2 < d1:
                return render_page("İzin Talep Et", "<div class='card'><div class='pill bad'>Bitiş tarihi başlangıçtan küçük olamaz.</div></div>", user=u)
        except Exception:
            return render_page("İzin Talep Et", "<div class='card'><div class='pill bad'>Tarih formatı YYYY-MM-DD olmalı.</div></div>", user=u)

        if leave_type == "Yillik":
            year = d1.year
            req_days = days_between_inclusive(start_date, end_date)
            remaining = remaining_annual_leave_days(u, year)
            if req_days > remaining:
                return render_page(
                    "İzin Talep Et",
                    f"<div class='card'><div class='pill bad'>Bu yıl kalan yıllık izniniz {remaining} gün. İstediğiniz {req_days} gün.</div></div>",
                    user=u
                )

        now = datetime.now().isoformat(timespec="seconds")

        # ✅ FIX: İlk aşama (sorumlu -> müdür -> owner)
        stage, pending_with = first_leave_stage_for_requester(u)

        is_half_day = 1 if (request.form.get("is_half_day") == "1") else 0
        half_day_part = (request.form.get("half_day_part") or "").strip()

        if is_half_day != 1:
            half_day_part = ""

        exec_sql("""
        INSERT INTO leave_requests (
            user_id,
            start_date,
            end_date,
            leave_type,
            reason,
            is_half_day,
            half_day_part,
            status,
            created_at,
            pending_with,
            stage
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            u["id"],
            start_date,
            end_date,
            leave_type,
            reason,
            is_half_day,
            half_day_part,
            STATUS_PENDING,
            now,
            pending_with,
            stage
        ))

        # Mail ilk onaycıya
        if stage == STAGE_RESPONSIBLE and pending_with:
            approver_email = get_user_email(pending_with)
        elif stage == STAGE_MANAGER and pending_with:
            approver_email = get_user_email(pending_with)
        else:
            approver_email = get_owner_email()

        approval_url = "https://ik.digiturkkibris.com/leave/admin"
        mail_html = ik_mail_template(
            title="Onay Bekleyen İzin Talebi",
            intro="Yeni bir izin talebi oluşturuldu. Detaylar aşağıdadır.",
            rows=[
                ("Talep Eden", f"{html_escape(u['full_name'])} ({html_escape(u['role'])})"),
                ("Tarih", f"{html_escape(start_date)} → {html_escape(end_date)}"),
                ("Tür", html_escape(leave_type)),
                ("Açıklama", nl2br(reason or "-")),
            ],
            stage_text=stage_label(stage),
            button_text="Onay / Red için Aç",
            button_url=approval_url,
            footer="İK Portal"
        )

        send_mail(
            approver_email,
            "[IK] Onay Bekleyen İzin Talebi",
            mail_html
        )

        return redirect("/leave/my")

    body = """
    <div class="card" style="max-width:720px;">
      <h1 class="h1">İzin Talep Et</h1>
      <form method="post">
        <div class="row">
          <div>
            <label class="muted">Başlangıç Tarihi</label>
            <input type="date" name="start_date" onchange="calculateDays()" required>
          </div>
          <div>
            <label class="muted">Bitiş Tarihi</label>
            <input type="date" name="end_date" onchange="calculateDays()" required>
          </div>
        </div>

        <div style="margin-top:12px">
          <label class="muted">Toplam Gün</label>
          <input id="totalDays" readonly placeholder="Tarih seçiniz">
        </div>

        <div class="row" style="margin-top:12px">
          <div>
            <label class="muted">İzin Türü</label>
            <select name="leave_type" id="leave_type" onchange="updateHalfDayAvailability()">
              <option>Yillik</option>
              <option>Rapor</option>
              <option>Mazeret</option>
              <option>Ucretsiz</option>
            </select>
          </div>

          <div>
            <label class="muted">İzin Süresi</label>
            <select name="is_half_day" id="is_half_day" onchange="toggleHalfDay()">
              <option value="0">Tam Gün</option>
              <option value="1">Yarım Gün</option>
            </select>
          </div>
        </div>

        <div id="halfDayPartBox" style="display:none;margin-top:12px">
          <label class="muted">Yarım Gün Bölümü</label>
          <select name="half_day_part">
            <option value="Sabah">Sabah</option>
            <option value="Öğleden Sonra">Öğleden Sonra</option>
          </select>
        </div>

      <script>

      function updateHalfDayAvailability() {
          var lt = document.getElementById("leave_type").value;
          var halfSel = document.getElementById("is_half_day");

          if (lt == "Yillik") {
              halfSel.disabled = false;
              toggleHalfDay();
          } else {
              halfSel.value = "0";
              halfSel.disabled = true;
              document.getElementById("halfDayPartBox").style.display = "none";
          }
      }

      document.addEventListener("DOMContentLoaded", function() {
          updateHalfDayAvailability();
      });

      function toggleHalfDay() {
          if (document.getElementById("is_half_day").disabled) {
              document.getElementById("halfDayPartBox").style.display = "none";
              return;
          }
          var val = document.getElementById("is_half_day").value;
          document.getElementById("halfDayPartBox").style.display = (val == "1") ? "block" : "none";
      }

      function calculateDays() {
          var start = document.querySelector("input[name='start_date']").value;
          var end = document.querySelector("input[name='end_date']").value;

          if (!start || !end) {
              document.getElementById("totalDays").value = "";
              return;
          }

          var d1 = new Date(start);
          var d2 = new Date(end);

          if (d2 < d1) {
              document.getElementById("totalDays").value = "Hatalı tarih";
              return;
          }

          var diffTime = d2 - d1;
          var diffDays = (diffTime / (1000 * 60 * 60 * 24)) + 1;

          document.getElementById("totalDays").value = diffDays + " gün";
      }
      </script>

        <div style="margin-top:12px">
          <label class="muted">Açıklama</label>
          <textarea name="reason" placeholder="Kısa açıklama..."></textarea>
        </div>
        <div style="margin-top:12px;display:flex;gap:10px;flex-wrap:wrap">
          <button class="btn" type="submit">Talep Gönder</button>
          <a class="btn2" href="/leave/my">İptal</a>
        </div>
      </form>
      <div class="note" style="margin-top:14px">
        <div class="muted">Talebiniz önce sorumlunuza (varsa), sonra müdüre (varsa), ardından patron/muhasebe onayına gider.</div>
      </div>
    </div>
    """
    return render_page("İzin Talep Et", body, user=u)

# -----------------------------
# Leave admin (manager/accounting/owner/responsible)
# -----------------------------
@app.route("/leave/admin")
@role_required(ROLE_OWNER, ROLE_ACCOUNTING, ROLE_MANAGER, ROLE_RESPONSIBLE)
def leave_admin():
    u = current_user()

    rows = query_all("""
    SELECT lr.*, us.full_name, us.role AS user_role
    FROM leave_requests lr
    JOIN users us ON us.id=lr.user_id
    ORDER BY lr.id DESC
    """)

    trs = ""
    for r in rows:
        stage = (r["stage"] or "").strip() or STAGE_OWNER
        pending_with = r["pending_with"]

        actions = ""
        if can_act_on_leave(u, r):
            actions = f"""
              <a class="btn2" href="/leave/admin/{r['id']}/approve">Onayla</a>
              <a class="btn2" href="/leave/admin/{r['id']}/reject">Reddet</a>
            """
        else:
            actions = "<span class='muted'>—</span>"

        stage_info = f"<span class='pill'>{html_escape(stage_label(stage))}</span>"
        if pending_with:
            pw = query_one("SELECT full_name FROM users WHERE id=?", (pending_with,))
            if pw:
                stage_info += f" <span class='muted'>({html_escape(pw['full_name'])})</span>"

        trs += f"""
        <tr>
          <td>{html_escape(r['full_name'])} <span class="pill">{html_escape(r['user_role'])}</span></td>
          <td>{html_escape(r['start_date'])}</td>
          <td>{html_escape(r['end_date'])}</td>
          <td>
            {html_escape(r['leave_type'])}
            { " (Yarım Gün - " + html_escape(r['half_day_part'] or "") + ")" if int(r['is_half_day'] or 0)==1 else "" }
          </td>
          <td>{format_status_pill(r['status'])}</td>
          <td>{stage_info}</td>
          <td>{actions}</td>
        </tr>
        """
    if not trs:
        trs = "<tr><td colspan='7' class='muted'>Kayıt yok.</td></tr>"

    body = f"""
    <div class="card">
      <h1 class="h1">İzin Yönetimi</h1>
      <div style="margin:10px 0 14px;display:flex;gap:10px;flex-wrap:wrap">
        <a class="btn2 btn-excel" href="/leave/admin.xlsx">Excel İndir</a>
      </div>
      <table>
        <thead><tr><th>Personel</th><th>Başlangıç</th><th>Bitiş</th><th>Tür</th><th>Durum</th><th>Aşama</th><th>İşlem</th></tr></thead>
        <tbody>{trs}</tbody>
      </table>
      <div class="note" style="margin-top:14px">
        <div class="muted">Sorumlu onayından sonra (müdür varsa) talep müdüre düşer. Final onay/ret sonrası personele mail gider.</div>
      </div>
    </div>
    """
    return render_page("İzin Yönetimi", body, user=u)

def _decide_leave(rid: int, action: str):
    """
    ✅ FIX: Çok aşamalı onay
    - Sorumlu ONAY -> (müdür varsa) müdüre gönder (status değişmez)
    - Müdür ONAY -> final ONAY
    - Owner/Muhasebe ONAY -> final ONAY
    - Her aşamada RED -> final RED
    """
    decider = current_user()
    lr = query_one("""
        SELECT lr.*, u.full_name, u.email, u.manager_id, u.responsible_id, u.role
        FROM leave_requests lr
        JOIN users u ON u.id = lr.user_id
        WHERE lr.id = ?
    """, (rid,))

    if not lr:
        abort(404)

    if lr["status"] != STATUS_PENDING:
        return

    if not can_act_on_leave(decider, lr):
        abort(403)

    stage = (lr["stage"] or "").strip() or STAGE_OWNER
    now = datetime.now().isoformat(timespec="seconds")

    if action == "reject":
        exec_sql("""
            UPDATE leave_requests
            SET status=?, decided_at=?, decided_by=?, pending_with=NULL, stage=?
            WHERE id=?
        """, (STATUS_REJECTED, now, decider["id"], STAGE_DONE, rid))

        requester_email = (lr["email"] or "").strip()
        if requester_email:
            mail_html = ik_mail_template(
                title="İzin Talebiniz REDDEDİLDİ",
                intro="İzin talebiniz reddedildi. Detaylar aşağıdadır.",
                rows=[
                    ("Tarih", f"{html_escape(lr['start_date'])} → {html_escape(lr['end_date'])}"),
                    ("Tür", html_escape(lr["leave_type"])),
                    ("Karar Veren", f"{html_escape(decider['full_name'])} ({html_escape(decider['role'])})"),
                    ("Tarih/Saat", html_escape(now)),
                ],
                stage_text="Sonuç: RED",
                button_text="İzinlerime Git",
                button_url="https://ik.digiturkkibris.com/leave/my",
                footer="İK Portal"
            )
            send_mail(requester_email, "[IK] İzin Talebiniz REDDEDİLDİ", mail_html)
        return

    # action == approve

    # Çakışma kontrolü (approve öncesi)
    conflict = query_one("""
        SELECT COUNT(*) AS cnt
        FROM leave_requests lr
        JOIN users u ON u.id = lr.user_id
        WHERE lr.status = ?
          AND u.role = ?
          AND lr.id != ?
          AND lr.start_date <= ?
          AND lr.end_date >= ?
    """, (
        STATUS_APPROVED,
        lr["role"],
        rid,
        lr["end_date"],
        lr["start_date"]
    ))

    if conflict and conflict["cnt"] >= 1:
        flash("⛔ Aynı rolde başka bir ONAYLI izin var. Onay işlemi durduruldu.", "error")
        return

    if stage == STAGE_RESPONSIBLE:
        requester_fake_row = {"manager_id": lr["manager_id"]}
        next_stage, next_pending = next_leave_stage_after_responsible(requester_fake_row)

        exec_sql("""
            UPDATE leave_requests
            SET pending_with=?, stage=?
            WHERE id=?
        """, (next_pending, next_stage, rid))

        if next_stage == STAGE_MANAGER and next_pending:
            to_mail = get_user_email(int(next_pending))
        else:
            to_mail = get_owner_email()

        approval_url = "https://ik.digiturkkibris.com/leave/admin"
        mail_html = ik_mail_template(
            title="Onay Bekleyen İzin Talebi",
            intro="Bir izin talebi bir sonraki aşamaya iletildi. Detaylar aşağıdadır.",
            rows=[
                ("Talep Eden", html_escape(lr["full_name"])),
                ("Tarih", f"{html_escape(lr['start_date'])} → {html_escape(lr['end_date'])}"),
                ("Tür", html_escape(lr["leave_type"])),
                ("Önceki Onay", f"{html_escape(decider['full_name'])} ({html_escape(decider['role'])})"),
            ],
            stage_text=f"Yeni Aşama: {stage_label(next_stage)}",
            button_text="Onay / Red için Aç",
            button_url=approval_url,
            footer="İK Portal"
        )
        send_mail(to_mail, "[IK] Onay Bekleyen İzin Talebi", mail_html)
        return

    if stage in (STAGE_MANAGER, STAGE_OWNER):
        exec_sql("""
            UPDATE leave_requests
            SET status=?, decided_at=?, decided_by=?, pending_with=NULL, stage=?
            WHERE id=?
        """, (STATUS_APPROVED, now, decider["id"], STAGE_DONE, rid))

        requester_email = (lr["email"] or "").strip()
        if requester_email:
            mail_html = ik_mail_template(
                title="İzin Talebiniz ONAYLANDI",
                intro="İzin talebiniz onaylandı. Detaylar aşağıdadır.",
                rows=[
                    ("Tarih", f"{html_escape(lr['start_date'])} → {html_escape(lr['end_date'])}"),
                    ("Tür", html_escape(lr["leave_type"])),
                    ("Karar Veren", f"{html_escape(decider['full_name'])} ({html_escape(decider['role'])})"),
                    ("Tarih/Saat", html_escape(now)),
                ],
                stage_text="Sonuç: ONAY",
                button_text="İzinlerime Git",
                button_url="https://ik.digiturkkibris.com/leave/my",
                footer="İK Portal"
            )
            send_mail(requester_email, "[IK] İzin Talebiniz ONAYLANDI", mail_html)
        return

@app.route("/leave/admin/<int:rid>/approve")
@role_required(ROLE_OWNER, ROLE_ACCOUNTING, ROLE_MANAGER, ROLE_RESPONSIBLE)
def leave_approve(rid):
    _decide_leave(rid, "approve")
    return redirect("/leave/admin")

@app.route("/leave/admin/<int:rid>/reject")
@role_required(ROLE_OWNER, ROLE_ACCOUNTING, ROLE_MANAGER, ROLE_RESPONSIBLE)
def leave_reject(rid):
    _decide_leave(rid, "reject")
    return redirect("/leave/admin")

@app.route("/leave/calendar")
@role_required(ROLE_OWNER, ROLE_ACCOUNTING, ROLE_MANAGER)
def leave_calendar():
    u = current_user()

    ym = (request.args.get("ym") or "").strip()
    today = date.today()

    if ym:
        try:
            parts = ym.split("-")
            year = int(parts[0])
            month = int(parts[1])
            if month < 1 or month > 12:
                raise ValueError("bad month")
        except Exception:
            year = today.year
            month = today.month
    else:
        year = today.year
        month = today.month

    first_day = date(year, month, 1)

    if month == 12:
        next_month_first = date(year + 1, 1, 1)
    else:
        next_month_first = date(year, month + 1, 1)

    last_day = next_month_first - timedelta(days=1)

    # Takvim ızgarası: Pazartesi başlangıç (0) -> Pazar bitiş (6)
    grid_start = first_day - timedelta(days=first_day.weekday())
    grid_end = last_day + timedelta(days=(6 - last_day.weekday()))

    grid_start_s = grid_start.isoformat()
    grid_end_s = grid_end.isoformat()

    role_filter = (request.args.get("role") or "").strip()

    role_sql = ""
    role_params = []

    if role_filter:
        role_sql = " AND us.role = ? "
        role_params.append(role_filter)

    base_sql = """
        SELECT lr.*, us.full_name, us.role AS user_role
        FROM leave_requests lr
        JOIN users us ON us.id = lr.user_id
        WHERE lr.status IN (?, ?)
          AND lr.start_date <= ?
          AND lr.end_date >= ?
    """

    params = [STATUS_PENDING, STATUS_APPROVED, grid_end_s, grid_start_s]

    # ✅ Manager sadece kendi ekibini görsün
    if u["role"] == ROLE_MANAGER:
        base_sql += " AND us.manager_id = ? "
        params.append(u["id"])

    # Rol filtresi (üstteki butonlar için)
    if role_filter:
        base_sql += " AND us.role = ? "
        params.append(role_filter)

    base_sql += " ORDER BY us.full_name ASC, lr.start_date ASC "

    rows = query_all(base_sql, tuple(params))

    day_map = {}
    for r in rows:
        try:
            s = datetime.strptime(r["start_date"], "%Y-%m-%d").date()
            e = datetime.strptime(r["end_date"], "%Y-%m-%d").date()
        except Exception:
            continue

        if s < grid_start:
            s = grid_start
        if e > grid_end:
            e = grid_end

        d = s
        while d <= e:
            k = d.isoformat()
            day_map.setdefault(k, []).append(r)
            d = d + timedelta(days=1)

    # Önceki / Sonraki ay linkleri
    if month == 1:
        prev_year, prev_month = year - 1, 12
    else:
        prev_year, prev_month = year, month - 1

    if month == 12:
        next_year, next_month = year + 1, 1
    else:
        next_year, next_month = year, month + 1

    prev_ym = f"{prev_year:04d}-{prev_month:02d}"
    next_ym = f"{next_year:04d}-{next_month:02d}"
    cur_ym = f"{year:04d}-{month:02d}"

    dow = ["Pzt", "Sal", "Çar", "Per", "Cum", "Cmt", "Paz"]

    header = f"""
    <div style="display:flex; align-items:center; justify-content:space-between; gap:10px; flex-wrap:wrap; margin-bottom:12px;">
      <div style="display:flex; gap:10px; align-items:center; flex-wrap:wrap;">
        <a class="btn2" href="/leave/calendar?ym={prev_ym}">← Önceki</a>
        <div style="font-weight:900; font-size:18px;">{html_escape(cur_ym)}</div>
        <a class="btn2" href="/leave/calendar?ym={next_ym}">Sonraki →</a>
      </div>
      <div style="display:flex; gap:8px; align-items:center; flex-wrap:wrap;">
        <span class="pill warn">Beklemede</span>
        <span class="pill ok">Onaylandi</span>
      </div>
      <div style="display:flex; gap:6px; flex-wrap:wrap; margin-top:6px;">
        <a class="btn2" href="/leave/calendar?ym={cur_ym}">Tümü</a>
        <a class="btn2" href="/leave/calendar?ym={cur_ym}&role={ROLE_MANAGER}">Müdürler</a>
        <a class="btn2" href="/leave/calendar?ym={cur_ym}&role={ROLE_RESPONSIBLE}">Sorumlular</a>
        <a class="btn2" href="/leave/calendar?ym={cur_ym}&role={ROLE_PERSONNEL}">Personel</a>
      </div>
    </div>
    """

    # CSS (sayfa içi)
    css = """
    <style>
      .calWrap {
        overflow-x: auto;
        -webkit-overflow-scrolling: touch;
      }

      .cal {
        display: grid;
        grid-template-columns: repeat(7, minmax(140px, 1fr));
        gap: 10px;
        min-width: 980px; /* 7 * 140 */
      }

      .calHead {
        font-weight: 900;
        color: #334155;
        text-align: center;
        padding: 8px 0;
        background: #f1f5f9;
        border-radius: 12px;
        border: 1px solid #e5e7eb;
      }

      .dayCell {
        border: 1px solid rgba(59,130,246,.20);
        border-radius: 14px;
        padding: 10px;
        min-height: 140px;
        background: rgba(30,41,59,.65);
        transition: all .15s ease;
      }

      .dayCell:hover {
        background: rgba(30,41,59,.85);
        border-color: rgba(59,130,246,.45);
        box-shadow: 0 0 0 1px rgba(59,130,246,.25);
      }

      .dayNum {
        font-weight: 900;
        display: flex;
        justify-content: space-between;
        align-items: center;
        margin-bottom: 6px;
      }

      .muted2 {
        color: #94a3b8;
        font-size: 12px;
      }

      .entry {
        display: flex;
        flex-direction: column;
        gap: 4px;
        padding: 8px;
        border-radius: 12px;
        background: rgba(15,23,42,.75);
        border: 1px solid rgba(59,130,246,.25);
        color: #e2e8f0;
      }

      .conflict {
        border: 2px solid #ef4444 !important;
        background: #fff1f2 !important;
      }

      .ann-item{padding:12px;border:1px solid rgba(148,163,184,.18);border-radius:14px;background:rgba(255,255,255,.02);transition:transform .12s ease, border-color .12s ease}
      .ann-item:hover{transform:translateY(-2px);border-color:rgba(99,102,241,.35)}
      .ann-top{display:flex;align-items:center;justify-content:space-between;gap:10px;flex-wrap:wrap}
      .ann-title a{color:inherit;text-decoration:none}
      .ann-title a:hover{text-decoration:underline}
      .ann-badges{display:flex;align-items:center;gap:8px}
      .ann-badge{font-size:11px;padding:4px 8px;border-radius:999px;border:1px solid rgba(148,163,184,.22);background:rgba(255,255,255,.04)}
      .ann-badge.new{
          background:linear-gradient(135deg,#ff3b3b,#ff0000);
          color:#ffffff;
          border:none;
          font-weight:700;
          box-shadow:0 0 8px rgba(255,0,0,.6);
          animation:pulseNew 1.4s infinite;
      }

      @keyframes pulseNew{
          0%{box-shadow:0 0 0 rgba(255,0,0,.7)}
          50%{box-shadow:0 0 14px rgba(255,0,0,.9)}
          100%{box-shadow:0 0 0 rgba(255,0,0,.7)}

      .ann-badges .ann-badge.new{
      background:linear-gradient(135deg,#ff3b3b,#ff0000) !important;
      color:#ffffff !important;
      border:0 !important;
      font-weight:900 !important;
      box-shadow:0 0 12px rgba(255,0,0,.85) !important;
      animation:pulseNew2 1.1s infinite !important;
      }

      @keyframes pulseNew2{
      0%{transform:scale(1);filter:brightness(1)}
      50%{transform:scale(1.08);filter:brightness(1.25)}
      100%{transform:scale(1);filter:brightness(1)}
      }

      }
    </style>
    """

    # Takvim body
    head_cells = "".join([f"<div class='calHead'>{x}</div>" for x in dow])

    day_cells = ""
    d = grid_start
    while d <= grid_end:
        k = d.isoformat()
        entries = day_map.get(k, [])
        is_other_month = (d.month != month)

        conflict_badge = ""
        cell_cls = "dayCell"
        if len(entries) >= 2:
            cell_cls += " conflict"
            conflict_badge = f"<span class='pill bad'>Çakışma: {len(entries)}</span>"

        inner = ""
        for r in entries:
            who = html_escape(r["full_name"])
            lt = html_escape(r["leave_type"])
            st = format_status_pill(r["status"])
            inner += f"""
            <div class="entry">
              <div style="font-weight:800; line-height:1.2;">{who}</div>
              <div class="muted2">{lt}</div>
              <div style="margin-top:4px;">{st}</div>
            </div>
            """

        day_cells += f"""
        <div class="{cell_cls}">
          <div class="dayNum">
            <div>{d.day}</div>
            <div>{conflict_badge}</div>
          </div>
          {("<div class='muted2'>Diğer ay</div>" if is_other_month else "")}
          {inner}
        </div>
        """

        d = d + timedelta(days=1)

    body = f"""
    <div class="card" style="max-width:1800px;margin:0 auto;">
      <h1 class="h1">İzin Takvimi</h1>
      {css}
      {header}
      <div class="calWrap">
        <div class="cal">
          {head_cells}
          {day_cells}
        </div>
      </div>
      <div class="note" style="margin-top:14px">
        <div class="muted">Aynı güne 2+ izin düşerse kırmızı “Çakışma” olarak işaretlenir. Beklemedeki talepler de görünür.</div>
      </div>
    </div>
    """

    return render_page("İzin Takvimi", body, user=u)

@app.route("/leave/admin.xlsx")
@role_required(ROLE_OWNER, ROLE_ACCOUNTING, ROLE_MANAGER, ROLE_RESPONSIBLE)
def leave_admin_xlsx():
    u = current_user()

    base_sql = """
        SELECT lr.*, us.full_name, us.username, us.role AS user_role
        FROM leave_requests lr
        JOIN users us ON us.id = lr.user_id
        WHERE 1=1
    """
    params = []

    # ✅ Manager sadece kendi ekibini görsün
    if u["role"] == ROLE_MANAGER:
        base_sql += " AND us.manager_id = ? "
        params.append(u["id"])

    # ✅ Responsible sadece kendi ekibini görsün
    if u["role"] == ROLE_RESPONSIBLE:
        base_sql += " AND us.responsible_id = ? "
        params.append(u["id"])

    base_sql += " ORDER BY lr.id DESC "

    rows = query_all(base_sql, tuple(params))

    wb = Workbook()
    ws = wb.active
    ws.title = "Izinler"

    ws.append([
        "ID",
        "Personel",
        "Kullanıcı",
        "Rol",
        "Başlangıç",
        "Bitiş",
        "Tür",
        "Yarım Gün",
        "Durum",
        "Aşama",
        "Oluşturma",
        "Karar Tarihi"
    ])

    for r in rows:
        ws.append([
            r["id"],
            r["full_name"],
            r["username"],
            r["user_role"],
            r["start_date"],
            r["end_date"],
            r["leave_type"],
            "Evet" if int(r["is_half_day"] or 0) == 1 else "Hayır",
            r["status"],
            r["stage"] or "",
            (r["created_at"] or "")[:19].replace("T", " "),
            (r["decided_at"] or "")[:19].replace("T", " ") if r["decided_at"] else ""
        ])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"izinler_{date.today().isoformat()}.xlsx"

    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


# -----------------------------
# ✅ ZIMMET YÖNETİMİ (NEW)
# -----------------------------
@app.route("/assets")
@role_required(ROLE_OWNER, ROLE_ACCOUNTING, ROLE_MANAGER, ROLE_RESPONSIBLE)
def assets_home():
    u = current_user()
    users = query_all("SELECT id, full_name, role FROM users WHERE is_active=1 ORDER BY full_name ASC")

    trs = ""
    for us in users:
        cnt = query_one("SELECT COUNT(*) AS c FROM assets WHERE user_id=? AND returned_at IS NULL", (us["id"],))["c"]
        trs += f"""
        <tr>
          <td>{html_escape(us['full_name'])}</td>
          <td><span class="pill">{html_escape(us['role'])}</span></td>
          <td class="muted">{cnt}</td>
          <td><a class="btn2" href="/assets/user/{us['id']}">Detay</a></td>
        </tr>
        """

    body = f"""
    <div class="card">
      <h1 class="h1">Zimmet Yönetimi</h1>
      <div style="margin:10px 0 14px;display:flex;gap:10px;flex-wrap:wrap">
        <a class="btn2 btn-excel" href="/assets.xlsx">Excel İndir</a>
      </div>
      <table>
        <thead><tr><th>Personel</th><th>Rol</th><th>Aktif Zimmet</th><th></th></tr></thead>
        <tbody>{trs}</tbody>
      </table>
      <div class="note" style="margin-top:14px">
        <div class="muted">Zimmet ekleme/çıkarma: Müdür / Muhasebe / Patron / Sorumlu yapabilir.</div>
      </div>
    </div>
    """
    return render_page("Zimmet Yönetimi", body, user=u)

@app.route("/assets.xlsx")
@role_required(ROLE_OWNER, ROLE_ACCOUNTING, ROLE_MANAGER, ROLE_RESPONSIBLE)
def assets_xlsx():
    u = current_user()

    base_sql = """
        SELECT a.*, us.full_name, us.username, us.role AS user_role,
               ub.full_name AS assigned_by_name,
               rb.full_name AS returned_by_name
        FROM assets a
        JOIN users us ON us.id = a.user_id
        LEFT JOIN users ub ON ub.id = a.assigned_by
        LEFT JOIN users rb ON rb.id = a.returned_by
        WHERE 1=1
    """
    params = []

    # ✅ Manager sadece kendi ekibini görsün
    if u["role"] == ROLE_MANAGER:
        base_sql += " AND us.manager_id = ? "
        params.append(u["id"])

    # ✅ Responsible sadece kendi ekibini görsün
    if u["role"] == ROLE_RESPONSIBLE:
        base_sql += " AND us.responsible_id = ? "
        params.append(u["id"])

    base_sql += " ORDER BY us.full_name ASC, a.id DESC "

    rows = query_all(base_sql, tuple(params))

    wb = Workbook()
    ws = wb.active
    ws.title = "Zimmetler"

    ws.append([
        "Zimmet ID",
        "Personel",
        "Kullanıcı",
        "Rol",
        "Ürün",
        "Özellik",
        "Seri No",
        "Zimmet Tarihi",
        "Zimmetleyen",
        "Durum",
        "Teslim Tarihi",
        "Teslim Alan",
        "Not"
    ])

    for r in rows:
        ws.append([
            r["id"],
            r["full_name"],
            r["username"],
            r["user_role"],
            r["item_name"],
            r["item_props"] or "",
            r["serial_no"] or "",
            (r["assigned_at"] or "")[:19].replace("T", " "),
            r["assigned_by_name"] or "",
            "Aktif" if not r["returned_at"] else "Teslim Edildi",
            (r["returned_at"] or "")[:19].replace("T", " ") if r["returned_at"] else "",
            r["returned_by_name"] or "",
            r["note"] or ""
        ])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"zimmetler_{date.today().isoformat()}.xlsx"

    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


@app.route("/assets/user/<int:uid>")
@role_required(ROLE_OWNER, ROLE_ACCOUNTING, ROLE_MANAGER, ROLE_RESPONSIBLE)
def assets_user(uid):
    u = current_user()
    us = query_one("SELECT id, full_name, role FROM users WHERE id=? AND is_active=1", (uid,))
    if not us:
        abort(404)

    rows = query_all("""
    SELECT a.*, ub.full_name AS assigned_by_name
    FROM assets a
    LEFT JOIN users ub ON ub.id=a.assigned_by
    WHERE a.user_id=?
    ORDER BY a.id DESC
    """, (uid,))

    trs = ""
    for r in rows:
        status = "<span class='pill ok'>Aktif</span>" if not r["returned_at"] else "<span class='pill warn'>Teslim Edildi</span>"
        returned = html_escape((r["returned_at"] or "")[:19].replace("T", " ")) if r["returned_at"] else "—"
        actions = ""
        if not r["returned_at"]:
            actions = f"<a class='btn2' href='/assets/{r['id']}/return'>Teslim Al</a>"
        else:
            actions = "<span class='muted'>—</span>"

        trs += f"""
        <tr>
          <td>{html_escape(r['item_name'])}</td>
          <td class="muted">{html_escape(r['item_props'] or '')}</td>
          <td class="muted">{html_escape(r['serial_no'] or '')}</td>
          <td class="muted">{html_escape((r['assigned_at'][:19].replace("T"," ")))}</td>
          <td class="muted">{html_escape(r['assigned_by_name'] or '')}</td>
          <td>{status}</td>
          <td class="muted">{returned}</td>
          <td>{actions}</td>
        </tr>
        """
    if not trs:
        trs = "<tr><td colspan='8' class='muted'>Kayıt yok.</td></tr>"

    body = f"""
    <div class="card">
      <h1 class="h1">{html_escape(us['full_name'])} - Zimmetler</h1>

      <div class="note" style="margin:12px 0;">
        <b>Yeni Zimmet Ekle</b>
        <form method="post" action="/assets/user/{uid}/add" style="margin-top:10px">
          <div class="row">
            <div>
              <label class="muted">İsim</label>
              <input name="item_name" placeholder="Laptop / Araba / Printer ..." required>
            </div>
            <div>
              <label class="muted">Seri No</label>
              <input name="serial_no" placeholder="SN12345">
            </div>
          </div>
          <div style="margin-top:10px">
            <label class="muted">Özellik</label>
            <input name="item_props" placeholder="Marka/Model/RAM vb...">
          </div>
          <div style="margin-top:10px">
            <label class="muted">Not</label>
            <input name="note" placeholder="İstersen not yaz...">
          </div>
          <div style="margin-top:10px;display:flex;gap:10px;flex-wrap:wrap">
            <button class="btn" type="submit">Zimmetle</button>
            <a class="btn2" href="/assets">Geri</a>
          </div>
        </form>
      </div>

      <table>
        <thead>
          <tr>
            <th>İsim</th><th>Özellik</th><th>Seri No</th><th>Zimmet Tarihi</th><th>Zimmetleyen</th><th>Durum</th><th>Teslim Tarihi</th><th></th>
          </tr>
        </thead>
        <tbody>{trs}</tbody>
      </table>
    </div>
    """
    return render_page("Zimmetler", body, user=u)

@app.route("/assets/user/<int:uid>/add", methods=["POST"])
@role_required(ROLE_OWNER, ROLE_ACCOUNTING, ROLE_MANAGER, ROLE_RESPONSIBLE)
def assets_add(uid):
    u = current_user()
    us = query_one("SELECT id FROM users WHERE id=? AND is_active=1", (uid,))
    if not us:
        abort(404)

    item_name = (request.form.get("item_name") or "").strip()
    item_props = (request.form.get("item_props") or "").strip()
    serial_no = (request.form.get("serial_no") or "").strip()
    note = (request.form.get("note") or "").strip()

    if not item_name:
        return redirect(f"/assets/user/{uid}")

    now = datetime.now().isoformat(timespec="seconds")
    exec_sql("""
    INSERT INTO assets (user_id, item_name, item_props, serial_no, assigned_at, assigned_by, note)
    VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (uid, item_name, item_props, serial_no, now, u["id"], note))

    return redirect(f"/assets/user/{uid}")

@app.route("/assets/<int:aid>/return")
@role_required(ROLE_OWNER, ROLE_ACCOUNTING, ROLE_MANAGER, ROLE_RESPONSIBLE)
def assets_return(aid):
    u = current_user()
    a = query_one("SELECT * FROM assets WHERE id=?", (aid,))
    if not a:
        abort(404)
    if a["returned_at"]:
        return redirect(f"/assets/user/{a['user_id']}")

    now = datetime.now().isoformat(timespec="seconds")
    exec_sql("""
    UPDATE assets
    SET returned_at=?, returned_by=?
    WHERE id=?
    """, (now, u["id"], aid))

    return redirect(f"/assets/user/{a['user_id']}")


# -----------------------------
# ✅ AVANS SİSTEMİ (NEW)
# -----------------------------
@app.route("/advance/my")
@login_required
def advance_my():
    u = current_user()
    if u["role"] == ROLE_OWNER:
        return render_page("Avanslarım", "<div class='card'><div class='pill warn'>Patron için avans talep ekranı kullanılmaz.</div></div>", user=u)

    rows = query_all("""
    SELECT * FROM advances
    WHERE user_id=?
    ORDER BY id DESC
    """, (u["id"],))

    trs = ""
    for r in rows:
        created = html_escape((r["created_at"][:19].replace("T"," ")))
        trs += f"""
        <tr>
          <td>{r['id']}</td>
          <td>{html_escape(str(r['amount']))}</td>
          <td class="muted">{html_escape(r['reason'] or '')}</td>
          <td>{format_status_pill(r['status'])}</td>
          <td class="muted">{created}</td>
        </tr>
        """
    if not trs:
        trs = "<tr><td colspan='5' class='muted'>Kayıt yok.</td></tr>"

    body = f"""
    <div class="card">
      <h1 class="h1">Avanslarım</h1>
      <div style="margin-bottom:12px;display:flex;gap:10px;flex-wrap:wrap">
        <a class="btn" href="/advance/new">Yeni Avans Talebi</a>
      </div>
      <table>
        <thead><tr><th>ID</th><th>Tutar</th><th>Açıklama</th><th>Durum</th><th>Oluşturma</th></tr></thead>
        <tbody>{trs}</tbody>
      </table>
      <div class="note" style="margin-top:14px">
        <div class="muted">Avans taleplerini sadece Muhasebe görür. Muhasebe isterse patron onayına gönderebilir yada kendi onaylayabilir.</div>
      </div>
    </div>
    """
    return render_page("Avanslarım", body, user=u)

@app.route("/advance/new", methods=["GET", "POST"])
@login_required
def advance_new():
    u = current_user()

    if u["role"] == ROLE_OWNER:
        return render_page("Avans Talep Et", "<div class='card'><div class='pill warn'>Patron için avans talep ekranı kapalıdır.</div></div>", user=u)

    if request.method == "POST":
        amount_raw = (request.form.get("amount") or "").strip().replace(",", ".")
        reason = (request.form.get("reason") or "").strip()

        try:
            amount = float(amount_raw)
        except Exception:
            return render_page("Avans Talep Et", "<div class='card'><div class='pill bad'>Tutar sayı olmalı.</div></div>", user=u)

        if amount <= 0:
            return render_page("Avans Talep Et", "<div class='card'><div class='pill bad'>Tutar 0'dan büyük olmalı.</div></div>", user=u)

        now = datetime.now().isoformat(timespec="seconds")
        exec_sql("""
        INSERT INTO advances (user_id, amount, reason, status, created_at)
        VALUES (?, ?, ?, ?, ?)
        """, (u["id"], amount, reason, ADV_STATUS_PENDING, now))

        acc_emails = get_accounting_emails()
        send_mail_many(
            acc_emails,
            "[IK] Yeni Avans Talebi",
            f"Talep Eden: {u['full_name']} ({u['role']})\n"
            f"Tutar: {amount}\n"
            f"Açıklama: {reason}\n\n"
            f"İncelemek için: https://ik.digiturkkibris.com/advance/accounting"
        )

        return redirect("/advance/my")

    body = """
    <div class="card" style="max-width:720px;">
      <h1 class="h1">Avans Talep Et</h1>
      <form method="post">
        <div class="row">
          <div>
            <label class="muted">Tutar</label>
            <input name="amount" placeholder="5000" required>
          </div>
          <div></div>
        </div>
        <div style="margin-top:12px">
          <label class="muted">Açıklama</label>
          <textarea name="reason" placeholder="Neden avans istiyorsunuz?"></textarea>
        </div>
        <div style="margin-top:12px;display:flex;gap:10px;flex-wrap:wrap">
          <button class="btn" type="submit">Talep Gönder</button>
          <a class="btn2" href="/advance/my">İptal</a>
        </div>
      </form>
      <div class="note" style="margin-top:14px">
        <div class="muted">Talep önce Muhasebe tarafından görülür.</div>
      </div>
    </div>
    """
    return render_page("Avans Talep Et", body, user=u)

@app.route("/advance/accounting")
@role_required(ROLE_ACCOUNTING)
def advance_accounting():
    u = current_user()
    rows = query_all("""
    SELECT a.*, us.full_name, us.role AS user_role
    FROM advances a
    JOIN users us ON us.id=a.user_id
    ORDER BY a.id DESC
    """)

    trs = ""
    for r in rows:
        actions = ""
        if r["status"] == ADV_STATUS_PENDING:
            actions = f"""
                <a class="btn2" href="/advance/{r['id']}/approve">Onayla</a>
                <a class="btn2" href="/advance/{r['id']}/reject">Reddet</a>
                <a class="btn2" href="/advance/{r['id']}/send_to_owner">Patrona Gönder</a>
            """
        elif r["status"] == ADV_STATUS_SENT_TO_OWNER:
            actions = "<span class='muted'>Patron onayında</span>"
        else:
            actions = "<span class='muted'>—</span>"

        trs += f"""
        <tr>
            <td>{r['id']}</td>
            <td>{html_escape(r['full_name'])} <span class="pill">{html_escape(r['user_role'])}</span></td>
            <td>{html_escape(str(r['amount']))}</td>
            <td class="muted">{html_escape(r['reason'] or '')}</td>
            <td>{format_status_pill(r['status'])}</td>
            <td>{actions}</td>
        </tr>
        """
    if not trs:
        trs = "<tr><td colspan='6' class='muted'>Kayıt yok.</td></tr>"

    body = f"""
    <div class="card">
      <h1 class="h1">Avans Yönetimi (Muhasebe)</h1>
      <div style="margin:10px 0 14px;display:flex;gap:10px;flex-wrap:wrap">
        <a class="btn2 btn-excel" href="/advance/accounting.xlsx">Excel İndir</a>
      </div>
      <table>
        <thead><tr><th>ID</th><th>Personel</th><th>Tutar</th><th>Açıklama</th><th>Durum</th><th>İşlem</th></tr></thead>
        <tbody>{trs}</tbody>
      </table>
      <div class="note" style="margin-top:14px">
        <div class="muted">Muhasebe isterse kendi onaylar/ret eder, isterse patron onayına gönderir.</div>
      </div>
    </div>
    """
    return render_page("Avans Yönetimi", body, user=u)

@app.route("/advance/accounting.xlsx")
@role_required(ROLE_ACCOUNTING)
def advance_accounting_xlsx():
    rows = query_all("""
    SELECT a.id, us.full_name, us.username, us.role AS user_role,
           a.amount, a.reason, a.status, a.created_at, a.decided_at
    FROM advances a
    JOIN users us ON us.id=a.user_id
    ORDER BY a.id DESC
    """)

    wb = Workbook()
    ws = wb.active
    ws.title = "Avanslar"

    ws.append([
        "ID", "Personel", "Kullanıcı", "Rol",
        "Tutar", "Açıklama", "Durum",
        "Oluşturma", "Karar Tarihi"
    ])

    for r in rows:
        ws.append([
            r["id"],
            r["full_name"],
            r["username"],
            r["user_role"],
            r["amount"],
            r["reason"] or "",
            r["status"],
            (r["created_at"] or "")[:19].replace("T", " "),
            (r["decided_at"] or "")[:19].replace("T", " ") if r["decided_at"] else ""
        ])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"avanslar_accounting_{date.today().isoformat()}.xlsx"
    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


@app.route("/advance/owner")
@role_required(ROLE_OWNER)
def advance_owner():
    u = current_user()
    rows = query_all("""
    SELECT a.*, us.full_name, us.role AS user_role
    FROM advances a
    JOIN users us ON us.id=a.user_id
    WHERE a.status=?
    ORDER BY a.id DESC
    """, (ADV_STATUS_SENT_TO_OWNER,))

    trs = ""
    for r in rows:
        trs += f"""
        <tr>
          <td>{r['id']}</td>
          <td>{html_escape(r['full_name'])} <span class="pill">{html_escape(r['user_role'])}</span></td>
          <td>{html_escape(str(r['amount']))}</td>
          <td class="muted">{html_escape(r['reason'] or '')}</td>
          <td>{format_status_pill(r['status'])}</td>
          <td>
            <a class="btn2" href="/advance/{r['id']}/owner_approve">Onayla</a>
            <a class="btn2" href="/advance/{r['id']}/owner_reject">Reddet</a>
          </td>
        </tr>
        """
    if not trs:
        trs = "<tr><td colspan='6' class='muted'>Patron onayında avans yok.</td></tr>"

    body = f"""
    <div class="card">
      <h1 class="h1">Avans Onay (Patron)</h1>
      <table>
        <thead><tr><th>ID</th><th>Personel</th><th>Tutar</th><th>Açıklama</th><th>Durum</th><th>İşlem</th></tr></thead>
        <tbody>{trs}</tbody>
      </table>
    </div>
    """
    return render_page("Avans Onay", body, user=u)

def _advance_decide(adv_id: int, new_status: str, decided_by_user):
    adv = query_one("""
    SELECT a.*, u.full_name, u.email
    FROM advances a
    JOIN users u ON u.id=a.user_id
    WHERE a.id=?
    """, (adv_id,))
    if not adv:
        abort(404)

    if new_status in (ADV_STATUS_APPROVED, ADV_STATUS_REJECTED):
        if adv["status"] not in (ADV_STATUS_PENDING, ADV_STATUS_SENT_TO_OWNER):
            return

    now = datetime.now().isoformat(timespec="seconds")
    exec_sql("""
    UPDATE advances
    SET status=?, decided_at=?, decided_by=?
    WHERE id=?
    """, (new_status, now, decided_by_user["id"], adv_id))

    requester_email = (adv["email"] or "").strip()
    if requester_email:
        subj = f"[IK] Avans Talebiniz {('ONAYLANDI' if new_status==ADV_STATUS_APPROVED else 'REDDEDİLDİ')}"
        body = (
            f"Merhaba,\n\n"
            f"Avans talebiniz sonuçlandı.\n\n"
            f"Durum: {new_status}\n"
            f"Tutar: {adv['amount']}\n"
            f"Açıklama: {adv['reason']}\n\n"
            f"Karar Veren: {decided_by_user['full_name']} ({decided_by_user['role']})\n"
            f"Tarih/Saat: {now}\n\n"
            f"İyi çalışmalar."
        )
        send_mail(requester_email, subj, body)

@app.route("/advance/<int:adv_id>/approve")
@role_required(ROLE_ACCOUNTING)
def advance_approve(adv_id):
    u = current_user()
    _advance_decide(adv_id, ADV_STATUS_APPROVED, u)
    return redirect("/advance/accounting")

@app.route("/advance/<int:adv_id>/reject")
@role_required(ROLE_ACCOUNTING)
def advance_reject(adv_id):
    u = current_user()
    _advance_decide(adv_id, ADV_STATUS_REJECTED, u)
    return redirect("/advance/accounting")

@app.route("/advance/<int:adv_id>/send_to_owner")
@role_required(ROLE_ACCOUNTING)
def advance_send_to_owner(adv_id):
    u = current_user()
    adv = query_one("""
    SELECT a.*, us.full_name, us.role AS user_role
    FROM advances a
    JOIN users us ON us.id=a.user_id
    WHERE a.id=?
    """, (adv_id,))
    if not adv:
        abort(404)

    if adv["status"] != ADV_STATUS_PENDING:
        return redirect("/advance/accounting")

    now = datetime.now().isoformat(timespec="seconds")
    exec_sql("""
    UPDATE advances
    SET status=?, forwarded_at=?, forwarded_by=?
    WHERE id=?
    """, (ADV_STATUS_SENT_TO_OWNER, now, u["id"], adv_id))

    owner_mail = get_owner_email()
    send_mail(
        owner_mail,
        "[IK] Patron Onayında Avans Talebi",
        f"Talep Eden: {adv['full_name']} ({adv['user_role']})\n"
        f"Tutar: {adv['amount']}\n"
        f"Açıklama: {adv['reason']}\n\n"
        f"Onay/Red için: https://ik.digiturkkibris.com/advance/owner"
    )

    return redirect("/advance/accounting")

@app.route("/advance/<int:adv_id>/owner_approve")
@role_required(ROLE_OWNER)
def advance_owner_approve(adv_id):
    u = current_user()
    _advance_decide(adv_id, ADV_STATUS_APPROVED, u)
    return redirect("/advance/owner")

@app.route("/advance/<int:adv_id>/owner_reject")
@role_required(ROLE_OWNER)
def advance_owner_reject(adv_id):
    u = current_user()
    _advance_decide(adv_id, ADV_STATUS_REJECTED, u)
    return redirect("/advance/owner")


# -----------------------------
# Users (owner only)
# -----------------------------
@app.route("/users")
@role_required(ROLE_OWNER, ROLE_ACCOUNTING)
def users_list():
    u = current_user()
    rows = query_all("""
        SELECT id, full_name, username, role, is_active, can_qr, email, manager_id, responsible_id, annual_leave_days, hire_date
        FROM users
        ORDER BY id DESC
    """)
    trs = ""
    for r in rows:
        qr = "Evet" if r["can_qr"] == 1 else "Hayır"
        active = "Aktif" if r["is_active"] == 1 else "Pasif"
        trs += f"""
        <tr>
          <td>{r['id']}</td>
          <td>{html_escape(r['full_name'])}</td>
          <td>{html_escape(r['username'])}</td>
          <td><span class="pill">{html_escape(r['role'])}</span></td>
          <td class="muted">{active}</td>
          <td class="muted">{html_escape(r['email'] or '')}</td>
          <td class="muted">{r['annual_leave_days']}</td>
          <td class="muted">{qr}</td>
          <td><a class="btn2" href="/users/{r['id']}/edit">Düzenle</a></td>
        </tr>
        """
    body = f"""
    <div class="card">
      <h1 class="h1">Kullanıcılar</h1>
      <div style="margin-bottom:12px">
        <a class="btn" href="/users/new">Yeni Kullanıcı</a>
      </div>
      <table>
        <thead><tr><th>ID</th><th>Ad Soyad</th><th>Kullanıcı</th><th>Rol</th><th>Durum</th><th>Email</th><th>Yıllık Hak</th><th>QR</th><th></th></tr></thead>
        <tbody>{trs}</tbody>
      </table>
      <div class="note" style="margin-top:14px">
        <div class="muted">Personel için müdür/sorumlu seçimi ve email tanımı yapmayı unutma (mail akışı için).</div>
      </div>
    </div>
    """
    return render_page("Kullanıcılar", body, user=u)

@app.route("/users/new", methods=["GET","POST"])
@role_required(ROLE_OWNER, ROLE_ACCOUNTING)
def users_new():
    u = current_user()

    managers = query_all("SELECT id, full_name FROM users WHERE role=? AND is_active=1 ORDER BY full_name", (ROLE_MANAGER,))
    responsibles = query_all("SELECT id, full_name FROM users WHERE role=? AND is_active=1 ORDER BY full_name", (ROLE_RESPONSIBLE,))

    if request.method == "POST":
        full_name = (request.form.get("full_name") or "").strip()
        username = (request.form.get("username") or "").strip().lower()
        password = (request.form.get("password") or "").strip()
        role = (request.form.get("role") or ROLE_PERSONNEL).strip()
        email = (request.form.get("email") or "").strip()
        hire_date = (request.form.get("hire_date") or "").strip()
        annual_leave_days = int((request.form.get("annual_leave_days") or "15").strip() or "15")
        annual_leave_override_raw = (request.form.get("annual_leave_override") or "").strip()
        annual_leave_override = int(annual_leave_override_raw) if annual_leave_override_raw.isdigit() else None
        manager_id = request.form.get("manager_id") or ""
        manager_id = int(manager_id) if manager_id.isdigit() else None

        responsible_id = request.form.get("responsible_id") or ""
        responsible_id = int(responsible_id) if responsible_id.isdigit() else None

        can_qr = 1 if (request.form.get("can_qr") == "1") else 0

        if not full_name or not username or not password:
            return render_page("Yeni Kullanıcı", "<div class='card'><div class='pill bad'>Zorunlu alanlar boş.</div></div>", user=u)
        if role not in ALLOWED_ROLES:
            role = ROLE_PERSONNEL

        qr_secret = secrets.token_hex(8) if can_qr else ""
        try:
            exec_sql("""
            INSERT INTO users (full_name, username, password_hash, role, is_active, email, manager_id, responsible_id, hire_date, annual_leave_days, annual_leave_override, can_qr, qr_secret)
            VALUES (?, ?, ?, ?, 1, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                full_name, username, generate_password_hash(password), role,
                email, manager_id, responsible_id, hire_date, annual_leave_days, annual_leave_override,
                can_qr, qr_secret
            ))
        except sqlite3.IntegrityError:
            return render_page("Yeni Kullanıcı", "<div class='card'><div class='pill bad'>Bu kullanıcı adı zaten var.</div></div>", user=u)

        return redirect("/users")

    mgr_opts = "<option value=''>—</option>" + "".join([f"<option value='{m['id']}'>{html_escape(m['full_name'])}</option>" for m in managers])
    resp_opts = "<option value=''>—</option>" + "".join([f"<option value='{m['id']}'>{html_escape(m['full_name'])}</option>" for m in responsibles])

    body = f"""
    <div class="card" style="max-width:900px">
      <h1 class="h1">Yeni Kullanıcı</h1>
      <form method="post">
        <div class="row">
          <div>
            <label class="muted">Ad Soyad</label>
            <input name="full_name" required>
          </div>
          <div>
            <label class="muted">Kullanıcı Adı</label>
            <input name="username" required>
          </div>
        </div>

        <div class="row" style="margin-top:12px">
          <div>
            <label class="muted">Şifre</label>
            <input name="password" type="password" required>
          </div>
          <div>
            <label class="muted">Rol</label>
            <select name="role">
              <option value="{ROLE_PERSONNEL}">personnel</option>
              <option value="{ROLE_MANAGER}">manager</option>
              <option value="{ROLE_RESPONSIBLE}">responsible</option>
              <option value="{ROLE_ACCOUNTING}">accounting</option>
              <option value="{ROLE_OWNER}">owner</option>
              <option value="{ROLE_IT}">it</option>
            </select>
          </div>
        </div>

        <div class="row" style="margin-top:12px">
          <div>
            <label class="muted">Email (bildirim için)</label>
            <input name="email" placeholder="personel@firma.com">
          </div>
          <div>
            <label class="muted">İşe Giriş Tarihi (YYYY-MM-DD)</label>
            <input name="hire_date" placeholder="2025-01-15">
          </div>
        </div>

        <div class="row" style="margin-top:12px">
          <div>
          <div style="margin-top:12px">
            <input type="hidden" name="annual_leave_days" value="15">
            <label class="muted">Manuel Yıllık Hak (Muhasebe Override)</label>
            <input name="annual_leave_override" placeholder="Boş bırakılırsa otomatik hesaplanır">
            <div class="muted" style="font-size:12px;margin-top:6px">
              Not: Doluysa otomatik kıdem hesabını ezer.
            </div>
          </div>

          </div>
          <div>
            <label class="muted">Müdür (Personel için)</label>
            <select name="manager_id">{mgr_opts}</select>
          </div>
        </div>

        <div class="row" style="margin-top:12px">
          <div>
            <label class="muted">Sorumlu (Personel için)</label>
            <select name="responsible_id">{resp_opts}</select>
          </div>
          <div></div>
        </div>

        <div style="margin-top:12px">
          <label><input type="checkbox" name="can_qr" value="1"> <b>QR ile Mesai kullanabilir</b></label>
        </div>

        <div style="margin-top:12px;display:flex;gap:10px;flex-wrap:wrap">
          <button class="btn" type="submit">Kaydet</button>
          <a class="btn2" href="/users">İptal</a>
        </div>
      </form>
    </div>
    """
    return render_page("Yeni Kullanıcı", body, user=u)

@app.route("/users/<int:uid>/edit", methods=["GET","POST"])
@role_required(ROLE_OWNER, ROLE_ACCOUNTING)
def users_edit(uid):
    u = current_user()
    row = query_one("SELECT * FROM users WHERE id=?", (uid,))
    if not row:
        abort(404)

    managers = query_all(
        "SELECT id, full_name FROM users WHERE role=? AND is_active=1 ORDER BY full_name",
        (ROLE_MANAGER,)
    )
    responsibles = query_all(
        "SELECT id, full_name FROM users WHERE role=? AND is_active=1 ORDER BY full_name",
        (ROLE_RESPONSIBLE,)
    )

    if request.method == "POST":
        # DEMO MODE: admin/owner kullanıcı değiştirilemez
        if is_demo_mode() and is_protected_user(row):
            return render_page(
                "Kullanıcı Düzenle",
                "<div class='card'><div class='pill bad'>DEMO MOD: Admin/Owner kullanıcı düzenlenemez.</div><div style='margin-top:10px'><a class='btn2' href='/users'>Geri</a></div></div>",
                user=u
            )        
        full_name = (request.form.get("full_name") or "").strip()
        role = (request.form.get("role") or ROLE_PERSONNEL).strip()
        is_active = 1 if (request.form.get("is_active") == "1") else 0
        can_qr = 1 if (request.form.get("can_qr") == "1") else 0

        email = (request.form.get("email") or "").strip()
        hire_date = (request.form.get("hire_date") or "").strip()

        # annual_leave_days zaten hidden olabilir ama yine de güvenli alalım
        annual_leave_days_raw = (request.form.get("annual_leave_days") or "15").strip()
        try:
            annual_leave_days = int(annual_leave_days_raw)
        except Exception:
            annual_leave_days = 15

        annual_leave_override_raw = (request.form.get("annual_leave_override") or "").strip()
        annual_leave_override = int(annual_leave_override_raw) if annual_leave_override_raw.isdigit() else None

        manager_id_raw = (request.form.get("manager_id") or "").strip()
        manager_id = int(manager_id_raw) if manager_id_raw.isdigit() else None

        responsible_id_raw = (request.form.get("responsible_id") or "").strip()
        responsible_id = int(responsible_id_raw) if responsible_id_raw.isdigit() else None

        new_pass = (request.form.get("new_pass") or "").strip()

        if not full_name:
            return render_page("Kullanıcı Düzenle", "<div class='card'><div class='pill bad'>Ad Soyad zorunlu.</div></div>", user=u)

        if role not in ALLOWED_ROLES:
            role = ROLE_PERSONNEL

        # QR secret yönetimi
        qr_secret = (row["qr_secret"] or "").strip()
        if can_qr == 1 and not qr_secret:
            qr_secret = secrets.token_hex(8)
        if can_qr != 1:
            # QR kapalıysa secret'ı temizlemek istersen aç:
            # qr_secret = ""
            pass

        if new_pass:
            exec_sql("""
                UPDATE users
                SET full_name=?,
                    role=?,
                    is_active=?,
                    email=?,
                    manager_id=?,
                    responsible_id=?,
                    hire_date=?,
                    annual_leave_days=?,
                    annual_leave_override=?,
                    can_qr=?,
                    qr_secret=?,
                    password_hash=?
                WHERE id=?
            """, (
                full_name,
                role,
                is_active,
                email,
                manager_id,
                responsible_id,
                hire_date,
                annual_leave_days,
                annual_leave_override,
                can_qr,
                qr_secret,
                generate_password_hash(new_pass),
                uid
            ))
        else:
            exec_sql("""
                UPDATE users
                SET full_name=?,
                    role=?,
                    is_active=?,
                    email=?,
                    manager_id=?,
                    responsible_id=?,
                    hire_date=?,
                    annual_leave_days=?,
                    annual_leave_override=?,
                    can_qr=?,
                    qr_secret=?
                WHERE id=?
            """, (
                full_name,
                role,
                is_active,
                email,
                manager_id,
                responsible_id,
                hire_date,
                annual_leave_days,
                annual_leave_override,
                can_qr,
                qr_secret,
                uid
            ))

        return redirect("/users")

    # GET: form
    mgr_opts = "<option value=''>—</option>" + "".join([
        f"<option value='{m['id']}' {'selected' if (row['manager_id']==m['id']) else ''}>{html_escape(m['full_name'])}</option>"
        for m in managers
    ])
    resp_opts = "<option value=''>—</option>" + "".join([
        f"<option value='{m['id']}' {'selected' if (row['responsible_id']==m['id']) else ''}>{html_escape(m['full_name'])}</option>"
        for m in responsibles
    ])

    checked_active = "checked" if int(row["is_active"] or 0) == 1 else ""
    checked_qr = "checked" if int(row["can_qr"] or 0) == 1 else ""

    body = f"""
    <div class="card" style="max-width:900px">
      <h1 class="h1">Kullanıcı Düzenle</h1>

      <form method="post">
        <div class="row">
          <div>
            <label class="muted">Ad Soyad</label>
            <input name="full_name" value="{html_escape(row['full_name'] or '')}" required>
          </div>
          <div>
            <label class="muted">Kullanıcı Adı</label>
            <input value="{html_escape(row['username'] or '')}" readonly>
          </div>
        </div>

        <div class="row" style="margin-top:12px">
          <div>
            <label class="muted">Rol</label>
            <select name="role">
              <option value="{ROLE_PERSONNEL}" {"selected" if row["role"]==ROLE_PERSONNEL else ""}>personnel</option>
              <option value="{ROLE_MANAGER}" {"selected" if row["role"]==ROLE_MANAGER else ""}>manager</option>
              <option value="{ROLE_RESPONSIBLE}" {"selected" if row["role"]==ROLE_RESPONSIBLE else ""}>responsible</option>
              <option value="{ROLE_ACCOUNTING}" {"selected" if row["role"]==ROLE_ACCOUNTING else ""}>accounting</option>
              <option value="{ROLE_OWNER}" {"selected" if row["role"]==ROLE_OWNER else ""}>owner</option>
              <option value="{ROLE_IT}" {"selected" if row["role"]==ROLE_IT else ""}>it</option>
            </select>
          </div>
          <div>
            <label class="muted">Yeni Şifre (opsiyonel)</label>
            <input name="new_pass" type="password" placeholder="Boş bırakılırsa değişmez">
          </div>
        </div>

        <div class="row" style="margin-top:12px">
          <div>
            <label class="muted">Email</label>
            <input name="email" value="{html_escape(row['email'] or '')}">
          </div>
          <div>
            <label class="muted">İşe Giriş Tarihi (YYYY-MM-DD)</label>
            <input name="hire_date" value="{html_escape(row['hire_date'] or '')}">
          </div>
        </div>

        <div class="row" style="margin-top:12px">
          <div>
            <input type="hidden" name="annual_leave_days" value="{int(row['annual_leave_days'] or 15)}">
            <label class="muted">Manuel Yıllık Hak (Override)</label>
            <input name="annual_leave_override" value="{html_escape('' if row['annual_leave_override'] is None else str(row['annual_leave_override']))}" placeholder="Boş bırakılırsa otomatik">
          </div>
          <div>
            <label class="muted">Müdür</label>
            <select name="manager_id">{mgr_opts}</select>
          </div>
        </div>

        <div class="row" style="margin-top:12px">
          <div>
            <label class="muted">Sorumlu</label>
            <select name="responsible_id">{resp_opts}</select>
          </div>
          <div>
            <label class="muted">Durum</label>
            <label style="display:flex;gap:8px;align-items:center;margin-top:10px">
              <input type="checkbox" name="is_active" value="1" {checked_active}>
              <span>Aktif</span>
            </label>
            <label style="display:flex;gap:8px;align-items:center;margin-top:10px">
              <input type="checkbox" name="can_qr" value="1" {checked_qr}>
              <span>QR ile Mesai kullanabilir</span>
            </label>
          </div>
        </div>

        <div style="margin-top:12px;display:flex;gap:10px;flex-wrap:wrap">
          <button class="btn" type="submit">Kaydet</button>
          <a class="btn2" href="/users">Geri</a>
        </div>
      </form>
    </div>
    """
    return render_page("Kullanıcı Düzenle", body, user=u)

# -----------------------------
# QR page (only can_qr)
# -----------------------------
@app.route("/qr")
@login_required
def qr_page():
    u = current_user()
    if u["can_qr"] != 1:
        abort(403)

    if not u["qr_secret"]:
        exec_sql("UPDATE users SET qr_secret=? WHERE id=?", (secrets.token_hex(8), u["id"]))
        u = current_user()

    token = make_token(u["qr_secret"], date.today())
    base = public_base_url()
    qr_url = f"{base}/giris/{token}"

    img = qrcode.make(qr_url)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    b64 = base64.b64encode(buf.getvalue()).decode("ascii")

    body = f"""
    <div class="card" style="max-width:760px;margin:0 auto;">
      <h1 class="h1">QR ile Mesai</h1>
      <p class="muted">Telefon kameranızla QR kodu okutun. Açılan sayfa işlemi tamamlayacak.</p>
      <div style="display:flex;justify-content:center;margin:16px 0;">
        <img src="data:image/png;base64,{b64}" style="width:320px;max-width:90%;border-radius:16px;border:1px solid rgba(15,23,42,.10)"/>
      </div>
      <div class="muted" style="text-align:center">
        Bugün için geçerlidir: <span class="pill">{date.today().isoformat()}</span>
      </div>
      <div style="margin-top:14px;text-align:center">
        <a class="btn2" href="/">Ana Sayfa</a>
      </div>
    </div>
    """
    return render_page("QR ile Mesai", body, user=u)

@app.route("/giris/<token>")
def qr_giris(token):
    d = date.today()
    u = find_user_by_token(token, d)
    if not u:
        body = """
        <div class="card" style="max-width:720px;margin:0 auto;">
          <h1 class="h1">Geçersiz QR</h1>
          <div class="pill bad">QR kod geçersiz ya da süresi dolmuş.</div>
        </div>
        """
        return render_page("QR", body, user=None)

    ip, ua = device_fingerprint()
    today = d.isoformat()

    other = device_used_by_another_user_today(today, ip, ua, u["id"])
    if other:
        body = f"""
        <div class="card" style="max-width:760px;margin:0 auto;">
          <h1 class="h1">Cihaz Kısıtı</h1>
          <div class="pill bad">Bu cihaz/IP bugün başka bir kullanıcı için kullanılmış.</div>
          <p class="muted" style="margin-top:10px">
            Mevcut kullanıcı: <b>{html_escape(other['full_name'])}</b>
          </p>
        </div>
        """
        return render_page("QR", body, user=None)

    # ✅ Art arda QR okutmayı engelle (cooldown)
    last_dt = last_attendance_dt(u["id"])
    if last_dt:
        diff_minutes = (datetime.now() - last_dt).total_seconds() / 60.0
        if diff_minutes < QR_COOLDOWN_MINUTES:
            kalan = int(QR_COOLDOWN_MINUTES - diff_minutes) + 1
            body = f"""
            <div class="card" style="max-width:760px;margin:0 auto;">
              <h1 class="h1">Çok Hızlı Okutma</h1>
              <div class="pill warn">Bu personel kısa süre önce QR okuttu.</div>
              <p class="muted" style="margin-top:10px">
                Lütfen <b>{kalan} dakika</b> sonra tekrar deneyin.
              </p>
              <p class="muted">{html_escape(u['full_name'])}</p>
            </div>
            """
            return render_page("QR", body, user=None)


    tur = next_attendance_tur(u["id"], today)
    if not tur:
        body = f"""
        <div class="card" style="max-width:760px;margin:0 auto;">
          <h1 class="h1">Tamamlandı</h1>
          <div class="pill warn">Bugün için tüm kayıtlar zaten girilmiş.</div>
          <p class="muted" style="margin-top:10px">{html_escape(u['full_name'])}</p>
        </div>
        """
        return render_page("QR", body, user=None)

    tarih, saat = insert_attendance(u["id"], tur, ip, ua)

    body = f"""
    <div class="card" style="max-width:760px;margin:0 auto;">
      <h1 class="h1">Kayıt Alındı ✅</h1>
      <div class="pill ok">{html_escape(u['full_name'])}</div>
      <div style="margin-top:12px" class="muted">
        Tarih: <b>{tarih}</b> &nbsp; Saat: <b>{saat}</b>
      </div>
      <div style="margin-top:10px">
        İşlem: <b>{html_escape(TUR_LABEL.get(tur, tur))}</b>
      </div>
    </div>
    """
    return render_page("QR", body, user=None)


# -----------------------------
# Attendance report (owner/accounting)
# -----------------------------
@app.route("/attendance", methods=["GET"])
@role_required(ROLE_OWNER, ROLE_ACCOUNTING)
def attendance_report():
    u = current_user()
    d1 = (request.args.get("d1") or (date.today() - timedelta(days=7)).isoformat()).strip()
    d2 = (request.args.get("d2") or date.today().isoformat()).strip()

    rows = query_all("""
    SELECT a.*, us.full_name, us.username
    FROM attendance a
    JOIN users us ON us.id=a.user_id
    WHERE a.tarih BETWEEN ? AND ?
    ORDER BY a.tarih DESC, us.full_name ASC, a.id ASC
    """, (d1, d2))

    trs = ""
    for r in rows:
        trs += f"""
        <tr>
          <td>{html_escape(r['tarih'])}</td>
          <td>{html_escape(r['saat'])}</td>
          <td>{html_escape(r['full_name'])}</td>
          <td>{html_escape(TUR_LABEL.get(r['tur'], r['tur']))}</td>
          <td class="muted">{html_escape((r['ip'] or '')[:24])}</td>
        </tr>
        """
    if not trs:
        trs = "<tr><td colspan='5' class='muted'>Kayıt yok.</td></tr>"

    body = f"""
    <div class="card">
      <h1 class="h1">Mesai Raporu</h1>
      <form method="get" class="row" style="margin:10px 0 14px">
        <div>
          <label class="muted">Başlangıç</label>
          <input name="d1" value="{html_escape(d1)}">
        </div>
        <div>
          <label class="muted">Bitiş</label>
          <input name="d2" value="{html_escape(d2)}">
        </div>
        <div style="grid-column:1/-1;display:flex;gap:10px;flex-wrap:wrap">
          <button class="btn" type="submit">Filtrele</button>
          <a class="btn2 btn-excel" href="/attendance.xlsx?d1={html_escape(d1)}&d2={html_escape(d2)}">Excel İndir</a>
        </div>
      </form>

      <table>
        <thead><tr><th>Tarih</th><th>Saat</th><th>Personel</th><th>İşlem</th><th>Cihaz/IP</th></tr></thead>
        <tbody>{trs}</tbody>
      </table>
    </div>
    """
    return render_page("Mesai Raporu", body, user=u)

@app.route("/attendance/team", methods=["GET"])
@role_required(ROLE_RESPONSIBLE)
def attendance_team_report():
    u = current_user()

    d1 = (request.args.get("d1") or (date.today() - timedelta(days=7)).isoformat()).strip()
    d2 = (request.args.get("d2") or date.today().isoformat()).strip()

    rows = query_all("""
    SELECT a.*, us.full_name, us.username
    FROM attendance a
    JOIN users us ON us.id=a.user_id
    WHERE us.responsible_id = ?
      AND a.tarih BETWEEN ? AND ?
    ORDER BY a.tarih DESC, us.full_name ASC, a.id ASC
    """, (u["id"], d1, d2))

    trs = ""
    for r in rows:
        trs += f"""
        <tr>
          <td>{html_escape(r['tarih'])}</td>
          <td>{html_escape(r['saat'])}</td>
          <td>{html_escape(r['full_name'])}</td>
          <td>{html_escape(TUR_LABEL.get(r['tur'], r['tur']))}</td>
          <td class="muted">{html_escape((r['ip'] or '')[:24])}</td>
        </tr>
        """
    if not trs:
        trs = "<tr><td colspan='5' class='muted'>Kayıt yok.</td></tr>"

    body = f"""
    <div class="card">
      <h1 class="h1">Mesai Raporu (Ekibim)</h1>

      <form method="get" class="row" style="margin:10px 0 14px">
        <div>
          <label class="muted">Başlangıç</label>
          <input name="d1" value="{html_escape(d1)}">
        </div>
        <div>
          <label class="muted">Bitiş</label>
          <input name="d2" value="{html_escape(d2)}">
        </div>
        <div style="grid-column:1/-1;display:flex;gap:10px;flex-wrap:wrap">
          <button class="btn" type="submit">Filtrele</button>
        </div>
      </form>

      <table>
        <thead><tr><th>Tarih</th><th>Saat</th><th>Personel</th><th>İşlem</th><th>Cihaz/IP</th></tr></thead>
        <tbody>{trs}</tbody>
      </table>
    </div>
    """
    return render_page("Mesai (Ekibim)", body, user=u)


@app.route("/attendance.xlsx")
@role_required(ROLE_OWNER, ROLE_ACCOUNTING)
def attendance_xlsx():
    d1 = (request.args.get("d1") or (date.today() - timedelta(days=7)).isoformat()).strip()
    d2 = (request.args.get("d2") or date.today().isoformat()).strip()

    rows = query_all("""
    SELECT a.tarih, a.saat, u.full_name, u.username, a.tur, a.ip, a.cihaz
    FROM attendance a
    JOIN users u ON u.id=a.user_id
    WHERE a.tarih BETWEEN ? AND ?
    ORDER BY a.tarih ASC, u.full_name ASC, a.id ASC
    """, (d1, d2))

    wb = Workbook()
    ws = wb.active
    ws.title = "Mesai"
    ws.append(["Tarih", "Saat", "Personel", "Kullanıcı", "İşlem", "IP", "Cihaz(User-Agent)"])

    for r in rows:
        ws.append([
            r["tarih"],
            r["saat"],
            r["full_name"],
            r["username"],
            TUR_LABEL.get(r["tur"], r["tur"]),
            r["ip"] or "",
            (r["cihaz"] or "")[:160]
        ])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"mesai_{d1}_to_{d2}.xlsx"
    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


# -----------------------------
# Startup
# -----------------------------

# ✅ Gunicorn için de DB'yi ayağa kaldır (import sırasında çalışır)
init_db()
ensure_default_owner()

@app.route("/api/dino/submit", methods=["POST"])
def dino_submit():
    if "user_id" not in session:
        return {"ok": False}, 403

    try:
        score = int(request.json.get("score", 0))
    except Exception:
        return {"ok": False}, 400

    if score <= 0:
        return {"ok": False}, 400

    exec_sql(
        "INSERT INTO dino_scores (user_id, score) VALUES (?, ?)",
        (session["user_id"], score)
    )

    return {"ok": True}


@app.route("/api/dino/leaderboard")
def dino_leaderboard():
    rows = query_all("""
        SELECT u.full_name, MAX(d.score) as best_score
        FROM dino_scores d
        JOIN users u ON u.id = d.user_id
        GROUP BY d.user_id
        ORDER BY best_score DESC
        LIMIT 10
    """)

    data = []
    for r in rows:
        data.append({
            "name": r["full_name"],
            "score": r["best_score"]
        })

    return {"leaders": data}

if __name__ == "__main__":
    app.run()
